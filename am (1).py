#!/usr/bin/env python3
"""
Image Analysis GUI Application (Scikit-Image Watershed Approach)
Tab 1: ROI Selection with Polygon Close-on-First-Click
Tab 2: Step-by-Step Watershed Segmentation using Sobel Gradient & Histogram Markers
Tab 3: Batch Image Processing with Indicator Classification
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from PIL import Image, ImageTk
import cv2
import numpy as np 
from scipy import ndimage as ndi
import json
import math
import os
import sys
import gc
import threading
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path


class ImageAnalyzerApp:
    def __init__(self, root):
        self.root = root
        #self.root.title("Indicator Detection")
        self.root.geometry("1400x900")
        self.root.configure(bg='#1e2328')

        # ── Colour palette  (professional dark-steel theme) ─────────────
        self.BG        = '#1e2328'   # near-black charcoal
        self.PANEL     = '#252b33'   # dark steel panel
        self.ACCENT    = '#2e3a4a'   # muted slate-blue accent
        self.HIGHLIGHT = '#2d7dd2'   # professional azure blue
        self.FG        = '#dde3ea'   # soft white text
        self.FG2       = '#7f8c9a'   # muted grey-blue subtext
        self.BTN_BG    = '#2e3a4a'   # slate button
        self.BTN_ACT   = '#1a5fa8'   # darker azure on hover
        self.ENTRY_BG  = '#181d22'   # deep input background

        # ── ttk style ───────────────────────────────────────────────────
        style = ttk.Style()
        style.theme_use('clam')

        style.configure('.',
                        background=self.BG,
                        foreground=self.FG,
                        fieldbackground=self.ENTRY_BG,
                        font=('Segoe UI', 10))

        style.configure('TFrame',       background=self.BG)
        style.configure('TLabel',       background=self.BG,  foreground=self.FG)
        style.configure('TLabelframe',  background=self.PANEL, foreground=self.FG)
        style.configure('TLabelframe.Label', background=self.PANEL, foreground=self.HIGHLIGHT,
                        font=('Segoe UI', 10, 'bold'))

        style.configure('TButton',
                        background=self.BTN_BG, foreground=self.FG,
                        font=('Segoe UI', 10, 'bold'), relief='flat',
                        borderwidth=0, padding=(10, 6))
        style.map('TButton',
                  background=[('active', self.BTN_ACT), ('disabled', '#2a3040')],
                  foreground=[('disabled', '#5a6472')])

        style.configure('Nav.TButton',
                        background=self.PANEL, foreground=self.FG2,
                        font=('Segoe UI', 11, 'bold'), relief='flat', padding=(16, 8))
        style.map('Nav.TButton',
                  background=[('active', self.HIGHLIGHT)],
                  foreground=[('active', '#ffffff')])

        style.configure('Active.Nav.TButton',
                        background=self.HIGHLIGHT, foreground='#ffffff',
                        font=('Segoe UI', 11, 'bold'), relief='flat', padding=(16, 8))

        style.configure('Danger.TButton',
                        background='#1e3a2f', foreground='#5dbf8e',
                        font=('Segoe UI', 10, 'bold'), relief='flat', padding=(10, 6))
        style.map('Danger.TButton', background=[('active', '#14532d')])

        style.configure('TEntry',
                        fieldbackground=self.ENTRY_BG, foreground=self.FG,
                        insertcolor=self.FG, relief='flat')
        style.configure('TScrollbar', background=self.ACCENT,
                        troughcolor=self.BG, bordercolor=self.BG)
        style.configure('Horizontal.TProgressbar',
                        troughcolor=self.PANEL, background=self.HIGHLIGHT,
                        bordercolor=self.BG)
        style.configure('TCheckbutton',
                        background=self.PANEL, foreground=self.FG)
        style.map('TCheckbutton', background=[('active', self.PANEL)])
        style.configure('TScale', background=self.PANEL, troughcolor=self.ACCENT)
        style.configure('TSpinbox',
                        fieldbackground=self.ENTRY_BG, foreground=self.FG,
                        background=self.ACCENT, arrowcolor=self.FG)
        style.configure('TSeparator', background=self.ACCENT)

        # ── Top header bar ──────────────────────────────────────────────
        header = tk.Frame(root, bg=self.PANEL, height=54)
        header.pack(side='top', fill='x')
        header.pack_propagate(False)

        # App title
        tk.Label(header, text='',
                 bg=self.PANEL, fg=self.HIGHLIGHT,
                 font=('Segoe UI', 14, 'bold')).pack(side='left', padx=20)

        # Logout button (right-aligned) – hidden until login
        self._logout_btn = tk.Button(header, text='Logout',
                               bg='#1e3a2f', fg='#5dbf8e', relief='flat',
                               font=('Segoe UI', 10, 'bold'), padx=14, pady=4,
                               activebackground='#14532d', activeforeground='#ffffff',
                               cursor='hand2',
                               command=self._logout)
        # Do NOT pack yet — shown after login

        # Navigation tabs (horizontal pill buttons)
        nav_frame = tk.Frame(header, bg=self.PANEL)
        nav_frame.pack(side='left', padx=30)

        self._tab_buttons = {}
        self._current_tab = None

        tabs = [
            ('single',  'Single Image'),
            ('batch',   'Batch Processing & 3D Visualization'),
        ]
        for key, label in tabs:
            btn = tk.Button(nav_frame, text=label,
                            bg=self.PANEL, fg=self.FG2, relief='flat',
                            font=('Segoe UI', 11, 'bold'), padx=16, pady=6,
                            activebackground=self.HIGHLIGHT, activeforeground='#ffffff',
                            cursor='hand2',
                            command=lambda k=key: self._switch_tab(k))
            btn.pack(side='left', padx=2)
            self._tab_buttons[key] = btn

        # Hide Single Image tab until login
        self._tab_buttons['single'].pack_forget()

        # ── Content area ────────────────────────────────────────────────
        self.content = tk.Frame(root, bg=self.BG)
        self.content.pack(fill='both', expand=True)

        self.tab_frames = {}
        for key, _ in tabs:
            f = tk.Frame(self.content, bg=self.BG)
            self.tab_frames[key] = f

        # ── Footer ──────────────────────────────────────────────────────
        footer = tk.Frame(root, bg=self.PANEL, height=28)
        footer.pack(side='bottom', fill='x')
        footer.pack_propagate(False)
        tk.Label(footer, text='Designed and Developed by IV',
                 bg=self.PANEL, fg=self.FG2, font=('Segoe UI', 9)).pack(expand=True)

        # ── Wire up legacy tab references ───────────────────────────────
        # Tab 1 = single image  (used by init_tab1 / watershed logic)
        self.tab2 = self.tab_frames['single']
        # Tab 3 = batch + 3D viz + reports (merged)
        self.tab3 = self.tab_frames['batch']

        # ── Initialise tab contents ─────────────────────────────────────
        self.init_tab1()          # ROI Selection  → popup window
        self.init_tab2()          # Watershed      → 'single' frame
        self._init_tab_batch_3d() # Merged Batch + 3D Viz + Reports → 'batch' frame

        # Show default tab
        self._switch_tab('batch')

        # ── Login state & Ctrl+A+B shortcut ────────────────────────────
        self._logged_in = False
        self._keys_down = set()
        self.root.bind('<KeyPress>',   self._track_key_down)
        self.root.bind('<KeyRelease>', self._track_key_up)

    # ── Shared GUI helpers ──────────────────────────────────────────────
    def _make_scrollable_card(self, parent, max_width=700):
        """Create a scrollable canvas with a centred card frame. Returns card frame."""
        outer = tk.Frame(parent, bg=self.BG)
        outer.pack(fill='both', expand=True)
        _canvas = tk.Canvas(outer, bg=self.BG, highlightthickness=0)
        _vsb = ttk.Scrollbar(outer, orient='vertical', command=_canvas.yview)
        _canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side='right', fill='y')
        _canvas.pack(side='left', fill='both', expand=True)
        card = tk.Frame(_canvas, bg=self.PANEL, bd=0, relief='flat')
        _card_id = _canvas.create_window((0, 0), window=card, anchor='n')
        def _on_card_resize(e):
            _canvas.configure(scrollregion=_canvas.bbox('all'))
        def _on_canvas_resize(e):
            w = min(max_width, e.width)
            _canvas.itemconfig(_card_id, width=w)
            _canvas.coords(_card_id, e.width // 2, 0)
            card.configure(width=w)
        card.bind('<Configure>', _on_card_resize)
        _canvas.bind('<Configure>', _on_canvas_resize)
        _canvas.bind_all('<MouseWheel>',
                         lambda e: _canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))
        return card

    def _make_browse_row(self, parent, label, var, pick_fn):
        """Shared label + readonly entry + Browse button row."""
        tk.Label(parent, text=label, bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=30)
        row = tk.Frame(parent, bg=self.PANEL)
        row.pack(fill='x', padx=30, pady=(2, 10))
        tk.Entry(row, textvariable=var, bg=self.ENTRY_BG, fg=self.FG2,
                 relief='flat', font=('Segoe UI', 9), state='readonly',
                 readonlybackground=self.ENTRY_BG, bd=0
                 ).pack(side='left', fill='x', expand=True, ipady=6, padx=(0, 8))
        tk.Button(row, text='Browse', command=pick_fn,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=10, pady=3,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

    def _make_btn(self, parent, text, cmd, primary=False, **kw):
        """Shared styled button factory. Caller can override any kwarg via **kw."""
        bg  = self.HIGHLIGHT if primary else self.BTN_BG
        act = self.BTN_ACT   if primary else self.HIGHLIGHT
        # Build defaults; caller's **kw takes precedence for any key
        defaults = dict(
            bg=bg, fg='#ffffff' if primary else self.FG,
            relief='flat', font=('Segoe UI', 10, 'bold'),
            padx=14, pady=6,
            activebackground=act, cursor='hand2',
        )
        defaults.update(kw)   # caller wins on any duplicate key
        return tk.Button(parent, text=text, command=cmd, **defaults)

    # ── Navigation helpers ──────────────────────────────────────────────
    def _switch_tab(self, key):
        # Single Image tab requires login
        if key == 'single' and not self._logged_in:
            self._open_login_popup()
            return
        for k, f in self.tab_frames.items():
            f.pack_forget()
        self.tab_frames[key].pack(fill='both', expand=True)
        for k, btn in self._tab_buttons.items():
            if k == key:
                btn.configure(bg=self.HIGHLIGHT, fg='#ffffff')
            else:
                btn.configure(bg=self.PANEL, fg=self.FG2)
        self._current_tab = key

    def _logout(self):
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self._logged_in = False
            # Hide Single Image tab button and logout button
            self._tab_buttons['single'].pack_forget()
            self._logout_btn.pack_forget()
            # Switch back to batch tab
            self._switch_tab('batch')

    # ── Ctrl+A+B key tracking ───────────────────────────────────────────
    def _track_key_down(self, event):
        self._keys_down.add(event.keysym.lower())
        ctrl_held = ('control_l' in self._keys_down or
                     'control_r' in self._keys_down or
                     'control'   in self._keys_down)
        if ctrl_held and 'a' in self._keys_down and 'b' in self._keys_down:
            self._keys_down.clear()          # prevent repeated triggers
            self._open_login_popup()

    def _track_key_up(self, event):
        self._keys_down.discard(event.keysym.lower())

    # ── Login popup (triggered by Ctrl+A+B or tab guard) ───────────────
    def _open_login_popup(self):
        """Show a modal login dialog over the main window."""
        # If already logged in just switch to single tab directly
        if self._logged_in:
            self._switch_tab_direct('single')
            return

        popup = tk.Toplevel(self.root)
        popup.title("Login")
        popup.resizable(False, False)
        popup.configure(bg=self.BG)
        popup.grab_set()        # modal

        # Centre over main window
        self.root.update_idletasks()
        rx = self.root.winfo_x() + (self.root.winfo_width()  - 340) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - 400) // 2
        popup.geometry(f"340x400+{rx}+{ry}")

        PANEL = self.PANEL; BLUE = self.HIGHLIGHT
        FG = self.FG; FG2 = self.FG2; ENTRY = self.ENTRY_BG

        card = tk.Frame(popup, bg=PANEL)
        card.place(relx=0.5, rely=0.5, anchor='center', width=300, height=360)

        tk.Label(card, text='⬡', bg=PANEL, fg=BLUE,
                 font=('Segoe UI', 32)).pack(pady=(28, 0))
        tk.Label(card, text='SINGLE IMAGE ACCESS', bg=PANEL, fg=FG,
                 font=('Segoe UI', 11, 'bold')).pack()
        tk.Label(card, text='Login required', bg=PANEL, fg=FG2,
                 font=('Segoe UI', 9)).pack(pady=(2, 20))

        tk.Label(card, text='Username', bg=PANEL, fg=FG2,
                 font=('Segoe UI', 9), anchor='w').pack(fill='x', padx=30)
        user_var = tk.StringVar()
        user_entry = tk.Entry(card, textvariable=user_var,
                              bg=ENTRY, fg=FG, relief='flat',
                              font=('Segoe UI', 11), insertbackground=FG, bd=0)
        user_entry.pack(fill='x', padx=30, ipady=7, pady=(3, 12))

        tk.Label(card, text='Password', bg=PANEL, fg=FG2,
                 font=('Segoe UI', 9), anchor='w').pack(fill='x', padx=30)
        pass_var = tk.StringVar()
        pass_entry = tk.Entry(card, textvariable=pass_var, show='●',
                              bg=ENTRY, fg=FG, relief='flat',
                              font=('Segoe UI', 11), insertbackground=FG, bd=0)
        pass_entry.pack(fill='x', padx=30, ipady=7, pady=(3, 6))

        err_lbl = tk.Label(card, text='', bg=PANEL, fg='#e05c5c',
                           font=('Segoe UI', 9))
        err_lbl.pack(pady=(0, 10))

        def _try():
            if (user_var.get().strip() == 'admin' and
                    pass_var.get().strip() == 'admin'):
                self._logged_in = True
                popup.destroy()
                # Reveal Single Image tab button and logout button
                self._tab_buttons['single'].pack(side='left', padx=2)
                self._logout_btn.pack(side='right', padx=16, pady=10)
                self._switch_tab_direct('single')
            else:
                err_lbl.config(text='Invalid username or password.')
                pass_var.set('')

        tk.Button(card, text='Login', command=_try,
                  bg=BLUE, fg='#ffffff', relief='flat',
                  font=('Segoe UI', 11, 'bold'), padx=20, pady=7,
                  activebackground=self.BTN_ACT, cursor='hand2').pack(
                  fill='x', padx=30)

        user_entry.bind('<Return>', lambda e: pass_entry.focus())
        pass_entry.bind('<Return>', lambda e: _try())
        user_entry.focus()

    def _switch_tab_direct(self, key):
        """Switch tab without the login gate (used after successful login)."""
        for k, f in self.tab_frames.items():
            f.pack_forget()
        self.tab_frames[key].pack(fill='both', expand=True)
        for k, btn in self._tab_buttons.items():
            btn.configure(bg=self.HIGHLIGHT if k == key else self.PANEL,
                          fg='#ffffff' if k == key else self.FG2)
        self._current_tab = key

    # ── 3D Visualise tab ────────────────────────────────────────────────
    def _init_tab_batch_3d(self):
        """Initialize the merged Batch Processing & 3D Visualization tab."""
        self.tab3_input_folder = None
        self.tab3_config_data  = None
        self.tab3_roi_points   = None
        self.tab3_processing   = False
        self._batch3d_stop_flag = False   # set True by Stop button

        # Reports state
        self._rpt_results_data = None
        self._rpt_metrics_data = None
        self._rpt_pred_folder  = None
        self._rpt_gt_folder    = None

        # 3D Viz hardcoded defaults (preserved from original)
        self._VIZ3D_SP_X   = 1.0
        self._VIZ3D_SP_Y   = 1.0
        self._VIZ3D_SP_Z   = 0.6 #10.0 0.48
        self._VIZ3D_ISLAND = 150
        self._VIZ3D_R_MIN  = 255
        self._VIZ3D_G_MAX  = 0
        self._VIZ3D_SX     = 2050.0
        self._VIZ3D_SY     = 2070.0
        self._VIZ3D_SZ     = 1360.0
        self._VIZ3D_TX     = 1030.0
        self._VIZ3D_TY     = 1050.0
        self._VIZ3D_TZ     = 670.0
        self._freecad_exe_var = tk.StringVar(
            value=r"/home/dlpda/AI_PROBLEM_STATEMENTS/Additive_Manufacturing/ws/prathyusha/squashfs-root/usr/bin/freecadcmd"
        )
         # ── Hardcoded 3D Slicer path ─────────────────────────────────────
        self._slicer_exe_var = tk.StringVar(
            value=r"/home/dlpda/manaswini/Defects_analysis_GUI/Slicer/Slicer"
        )

        frame = self.tab_frames['batch']
        frame.configure(bg=self.BG)

        # ── Helper ───────────────────────────────────────────────────────
        def _row_entry(parent, label_text, var, cmd, btn_text='Browse'):
            row = tk.Frame(parent, bg=self.PANEL)
            row.pack(fill='x', padx=16, pady=5)
            tk.Label(row, text=label_text, bg=self.PANEL, fg=self.FG2,
                     font=('Segoe UI', 9), width=16, anchor='w').pack(side='left')
            entry = tk.Entry(row, textvariable=var, bg=self.ENTRY_BG, fg=self.FG,
                             relief='flat', font=('Segoe UI', 9), state='readonly',
                             readonlybackground=self.ENTRY_BG, bd=0)
            entry.pack(side='left', fill='x', expand=True, ipady=6, padx=(6, 8))
            tk.Button(row, text=btn_text, command=cmd,
                      bg=self.BTN_BG, fg=self.FG, relief='flat',
                      font=('Segoe UI', 9, 'bold'), padx=10,
                      activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

        # ── Settings card ────────────────────────────────────────────────
        card = tk.Frame(frame, bg=self.PANEL)
        card.pack(fill='x', padx=20, pady=(18, 0))

        tk.Label(card, text='',
                 bg=self.PANEL, fg=self.HIGHLIGHT,
                 font=('Segoe UI', 13, 'bold')).pack(anchor='w', padx=16, pady=(14, 10))

        # Input Folder row
        self._tab3_folder_var = tk.StringVar(value='Not selected')
        _row_entry(card, 'Input Folder :', self._tab3_folder_var,
                   self._batch3d_pick_input_folder)

        # STL File row
        self._batch3d_stl_var = tk.StringVar(value='Not selected')
        def _pick_stl():
            p = filedialog.askopenfilename(
                title='Select STL File',
                filetypes=[('STL / STEP files', '*.stl *.step *.stp'),
                           ('All files', '*.*')])
            if p:
                self._batch3d_stl_var.set(p)
        _row_entry(card, 'STL File :', self._batch3d_stl_var, _pick_stl)

        # Output Folder row
        self._tab3_output_var = tk.StringVar(value='Not selected')
        def _pick_out():
            p = filedialog.askdirectory(title='Select Output Folder')
            if p:
                self._tab3_output_var.set(p)
        _row_entry(card, 'Output Folder :', self._tab3_output_var, _pick_out)

        # ── Separator ────────────────────────────────────────────────────
        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=16, pady=(8, 10))

        # ── Action buttons ────────────────────────────────────────────────
        btn_row = tk.Frame(card, bg=self.PANEL)
        btn_row.pack(fill='x', padx=16, pady=(0, 14))

        self._batch3d_start_btn = tk.Button(
            btn_row, text='  Start',
            command=self._batch3d_start,
            bg=self.HIGHLIGHT, fg='#ffffff', relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=18, pady=6,
            activebackground=self.BTN_ACT, cursor='hand2')
        self._batch3d_start_btn.pack(side='left', padx=(0, 6))

        self._batch3d_stop_btn = tk.Button(
            btn_row, text=' Stop',
            command=self._batch3d_stop,
            bg=self.BTN_BG, fg=self.FG, relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=18, pady=6,
            activebackground=self.HIGHLIGHT, cursor='hand2')
        self._batch3d_stop_btn.pack(side='left', padx=(0, 6))

        self._batch3d_clear_btn = tk.Button(
            btn_row, text=' Clear',
            command=self._batch3d_clear,
            bg=self.BTN_BG, fg=self.FG, relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=18, pady=6,
            activebackground=self.HIGHLIGHT, cursor='hand2')
        self._batch3d_clear_btn.pack(side='left', padx=(0, 6))

        self._batch3d_viz_btn = tk.Button(
            btn_row, text=' 3D Visualise',
            command=self._batch3d_launch_viz_independent,
            bg='#1e3a2f', fg='#5dbf8e', relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=18, pady=6,
            activebackground='#14532d', cursor='hand2')
        self._batch3d_viz_btn.pack(side='left', padx=(0, 6))

        self._batch3d_reports_btn = tk.Button(
            btn_row, text=' Reports',
            command=self._open_reports_popup,
            bg=self.BTN_BG, fg=self.FG, relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=18, pady=6,
            activebackground=self.HIGHLIGHT, cursor='hand2')
        self._batch3d_reports_btn.pack(side='left')

        # ── Change 3: clickable report hyperlink (not a separate button) ──
        # Always points to the most recently generated report; styled and
        # behaves like a real hyperlink (underlined link-colored text,
        # hand cursor) rather than a button.
        self._batch3d_report_link_btn = tk.Label(
            btn_row, text='No report generated yet',
            bg=self.PANEL, fg=self.FG2,
            font=('Segoe UI', 9, 'underline'), cursor='arrow')
        self._batch3d_report_link_btn.pack(side='left', padx=(14, 0))
        self._batch3d_report_link_btn.bind('<Button-1>', lambda e: self._batch3d_open_report())
        self._latest_report_path = None

        # ── Progress card ─────────────────────────────────────────────────
        prog_card = tk.Frame(frame, bg=self.PANEL)
        prog_card.pack(fill='x', padx=20, pady=(10, 0))

        prog_top = tk.Frame(prog_card, bg=self.PANEL)
        prog_top.pack(fill='x', padx=16, pady=(12, 4))

        self._batch3d_status_lbl = tk.Label(
            prog_top, text='Ready', bg=self.PANEL, fg=self.HIGHLIGHT,
            font=('Segoe UI', 10, 'bold'))
        self._batch3d_status_lbl.pack(side='left')

        self._batch3d_progress = ttk.Progressbar(
            prog_top, orient='horizontal', mode='determinate', length=500)
        self._batch3d_progress.pack(side='right', fill='x', expand=True, padx=(16, 8))

        self._batch3d_pct_lbl = tk.Label(
            prog_top, text='0%', bg=self.PANEL, fg=self.FG2,
            font=('Segoe UI', 9, 'bold'), width=5, anchor='e')
        self._batch3d_pct_lbl.pack(side='right')

        
        # ── Status / Log card ─────────────────────────────────────────────
        log_card = tk.Frame(frame, bg=self.PANEL)
        log_card.pack(fill='both', expand=True, padx=20, pady=(10, 12))

        tk.Label(log_card, text='Status', bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=16, pady=(10, 4))

        log_outer = tk.Frame(log_card, bg=self.BG)
        log_outer.pack(fill='both', expand=True, padx=10, pady=(0, 10))

        log_sb = ttk.Scrollbar(log_outer)
        log_sb.pack(side='right', fill='y')

        self.tab3_log_text = scrolledtext.ScrolledText(
            log_outer, height=18,
            bg='#181d22', fg='#b0bec9',
            font=('Consolas', 9), relief='flat', bd=0,
            insertbackground=self.FG,
            yscrollcommand=log_sb.set)
        self.tab3_log_text.pack(fill='both', expand=True)
        log_sb.config(command=self.tab3_log_text.yview)

        # ── Auto-load conf.json ───────────────────────────────────────────
        _conf_path = Path(__file__).resolve().parent / 'configurations' / 'conf.json'
        if _conf_path.exists():
            try:
                with open(_conf_path, 'r') as _f:
                    _conf = json.load(_f)
                if 'polygon_points' in _conf or 'pipeline_params' in _conf:
                    self.tab3_config_data = _conf
                    if 'polygon_points' in _conf:
                        self.tab3_roi_points = np.array(_conf['polygon_points'], dtype=np.int32)
                    self.tab3_log(f"Auto-loaded configuration: {_conf_path}")
                else:
                    self.tab3_log("conf.json found but missing required keys.")
            except Exception as _e:
                self.tab3_log(f"Could not load conf.json: {_e}")

    # ── Merged tab helpers ───────────────────────────────────────────────

    def _batch3d_pick_input_folder(self):
        """Pick input folder and auto-load config.json from Configurations/ subfolder."""
        folder_path = filedialog.askdirectory(title='Select Input Images Folder')
        if not folder_path:
            return
        self.tab3_input_folder = folder_path
        self._tab3_folder_var.set(folder_path)
        self.tab3_log(f"Input folder selected: {folder_path}")

        # Auto-load Configurations/config.json relative to the input folder
        conf_candidates = [
            os.path.join(folder_path, 'Configurations', 'config.json'),
            os.path.join(folder_path, '..', 'Configurations', 'config.json'),
            str(Path(__file__).resolve().parent / 'configurations' / 'conf.json'),
        ]
        loaded = False
        for conf_path in conf_candidates:
            conf_path = os.path.normpath(conf_path)
            if os.path.isfile(conf_path):
                try:
                    with open(conf_path, 'r') as f:
                        data = json.load(f)
                    self.tab3_config_data = data
                    if 'polygon_points' in data:
                        self.tab3_roi_points = np.array(data['polygon_points'], dtype=np.int32)
                    self.tab3_log(f"Configuration loaded successfully from: {conf_path}")
                    loaded = True
                    break
                except Exception as e:
                    self.tab3_log(f"Could not load config from {conf_path}: {e}")
        if not loaded:
            self.tab3_log("Configuration file not found — will use defaults.")

    def _batch3d_stop(self):
        """Signal running task to stop."""
        self._batch3d_stop_flag = True
        self.tab3_processing = False
        self.tab3_log("Operation cancelled by user.")
        self._batch3d_status_lbl.config(text='Stopped')

    def _batch3d_clear(self):
        """Clear all UI inputs and status without deleting output files."""
        self._tab3_folder_var.set('Not selected')
        self._batch3d_stl_var.set('Not selected')
        self._tab3_output_var.set('Not selected')
        self.tab3_input_folder = None
        self._batch3d_set_progress(0)
        self._batch3d_status_lbl.config(text='Ready')
        self.tab3_log_text.delete('1.0', tk.END)
        self._batch3d_report_link_btn.config(
            text='No report generated yet', fg=self.FG2, cursor='arrow')
        self._latest_report_path = None
        self.tab3_log("UI cleared successfully.")

    def _batch3d_start(self):
        """Start button: batch → report → 3D viz (auto chain)."""
        if self.tab3_processing:
            return
        if not self.tab3_input_folder or self.tab3_input_folder == 'Not selected':
            messagebox.showwarning('Start', 'Please select an Input Folder first.')
            return
        output_base = self._tab3_output_var.get()
        if not output_base or output_base == 'Not selected':
            messagebox.showwarning('Start', 'Please select an Output Folder.')
            return

        self._batch3d_stop_flag = False
        self.tab3_processing = True
        self._batch3d_start_btn.config(state='disabled')
        self._batch3d_set_progress(0)
        self._batch3d_status_lbl.config(text='Batch Processing…')

        # Collect images
        image_extensions = ['.png', '.jpg', '.jpeg']
        image_files = [
            os.path.join(self.tab3_input_folder, f)
            for f in os.listdir(self.tab3_input_folder)
            if os.path.isfile(os.path.join(self.tab3_input_folder, f))
            and any(f.lower().endswith(ext) for ext in image_extensions)
        ]
        if not image_files:
            messagebox.showwarning('Start', 'No image files found in selected folder.')
            self.tab3_processing = False
            self._batch3d_start_btn.config(state='normal')
            return

        self.tab3_log(f"\nFound {len(image_files)} image(s) to process")

        defects_path     = os.path.join(output_base, 'defects')
        non_defects_path = os.path.join(output_base, 'non_defects')
        os.makedirs(defects_path, exist_ok=True)
        os.makedirs(non_defects_path, exist_ok=True)
        self.tab3_log(f"Output → Indicators: {defects_path}")
        self.tab3_log(f"Output → Non Indicators: {non_defects_path}")

        def _run():
            total        = len(image_files)
            defect_count = 0
            non_defect   = 0
            segment_counts = {}   # filename -> number of segments (reused by report)
            import multiprocessing as _mp
            workers = max(1, min(_mp.cpu_count(), 32))
            self.root.after(0, lambda: self.tab3_log(
                f"Starting parallel processing with {workers} worker(s)…"))

            from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed
            futures = {}
            UPDATE_EVERY = max(1, total // 20)

            with ThreadPoolExecutor(max_workers=workers) as executor:
                for img_path in image_files:
                    if self._batch3d_stop_flag:
                        break
                    fut = executor.submit(
                        self.tab3_process_single_image,
                        img_path, defects_path, non_defects_path, img_path)
                    futures[fut] = img_path

                for done_idx, fut in enumerate(_as_completed(futures), start=1):
                    if self._batch3d_stop_flag:
                        break
                    img_path = futures[fut]
                    fname = os.path.basename(img_path)
                    try:
                        num_seg = fut.result()
                        segment_counts[fname] = num_seg
                        if num_seg >= 1:
                            defect_count += 1
                            msg = f"✓ {fname} → Indicator ({num_seg} segments)"
                        else:
                            non_defect += 1
                            msg = f"✓ {fname} → Non Indicator"
                    except Exception as exc:
                        segment_counts[fname] = 0
                        msg = f"✗ {fname} → ERROR: {exc}"
                    del fut
                    gc.collect()

                    # Progress: batch phase is 0-70% — update on every image
                    # so the bar/percentage advance smoothly without large jumps.
                    progress = (done_idx / total) * 70
                    show_log = (done_idx % UPDATE_EVERY == 0 or done_idx == total)
                    _msg = msg; _di = done_idx; _log_it = show_log
                    def _update(p=progress, m=_msg, di=_di, log_it=_log_it):
                        self._batch3d_set_progress(p)
                        self._batch3d_status_lbl.config(text=f"Processing {di}/{total}…")
                        if log_it:
                            self.tab3_log(m)
                    self.root.after(0, _update)

            if self._batch3d_stop_flag:
                def _stopped():
                    self.tab3_processing = False
                    self._batch3d_start_btn.config(state='normal')
                self.root.after(0, _stopped)
                return

            # ── Phase 2: Report generation (70→75%) ─────────────────────
            def _do_report():
                self._batch3d_set_progress(70)
                self._batch3d_status_lbl.config(text='Generating report…')
                self.tab3_log("\nGenerating report…")
                report_path = os.path.join(output_base, 'report.xlsx')
                try:
                    self._batch3d_generate_report_auto(
                        output_base, defects_path, non_defects_path, report_path,
                        segment_counts=segment_counts)
                    self.tab3_log(f"✓ Report generated → {report_path}")
                    self._batch3d_set_report_link(report_path)
                except Exception as e:
                    self.tab3_log(f"⚠ Report generation failed: {e}")
                self._batch3d_set_progress(75)
                self._batch3d_status_lbl.config(text='Report done.')
                self.tab3_log(f"\nBatch complete — Indicators: {defect_count} | Non Indicators: {non_defect}")

                # ── Phase 3: Auto 3D Viz (75→100%) ──────────────────────
                stl_path = self._batch3d_stl_var.get()
                if stl_path and stl_path != 'Not selected' and os.path.isfile(stl_path):
                    self.tab3_log("\n3D Visualization starting automatically…")
                    self._batch3d_status_lbl.config(text='3D Visualization…')
                    self._batch3d_run_viz(output_base, stl_path, auto=True)
                else:
                    self.tab3_log("STL file not selected — skipping auto 3D Visualization.")
                    self._batch3d_set_progress(100)
                    self._batch3d_status_lbl.config(text='Complete!')

                self.tab3_processing = False
                self._batch3d_start_btn.config(state='normal')

            self.root.after(0, _do_report)

        threading.Thread(target=_run, daemon=True).start()

    def _batch3d_set_progress(self, value):
        """Update the progress bar AND its percentage label together (Change 1)."""
        value = max(0, min(100, value))
        self._batch3d_progress['value'] = value
        if hasattr(self, '_batch3d_pct_lbl'):
            self._batch3d_pct_lbl.config(text=f"{int(round(value))}%")

    def _batch3d_set_report_link(self, report_path):
        """Enable/refresh the clickable Report hyperlink (Change 3). Always
        points to the most recently generated report; stays clickable until
        replaced by a newer report."""
        self._latest_report_path = report_path
        if hasattr(self, '_batch3d_report_link_btn'):
            self._batch3d_report_link_btn.config(
                text=f'📄  Open Report ({os.path.basename(report_path)})',
                fg=self.HIGHLIGHT, cursor='hand2')

    def _batch3d_open_report(self):
        """Open the latest generated report with the OS-default application."""
        path = getattr(self, '_latest_report_path', None)
        if not path or not os.path.isfile(path):
            messagebox.showwarning('Report', 'No report has been generated yet.')
            return
        try:
            if sys.platform.startswith('win'):
                os.startfile(path)
            elif sys.platform == 'darwin':
                import subprocess as _sp
                _sp.Popen(['open', path])
            else:
                import subprocess as _sp
                _sp.Popen(['xdg-open', path])
        except Exception as e:
            messagebox.showerror('Report', f'Could not open report:\n{e}')

    def _batch3d_generate_report_auto(self, output_base, defects_path, non_defects_path, report_path,
                                       segment_counts=None):
        """Auto-generate a simple summary xlsx after batch processing.

        segment_counts: optional {filename: num_segments} dict captured during
        batch processing (Change 5) — reused here rather than recomputed.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        segment_counts = segment_counts or {}

        # Collect image names from each folder
        def _list_images(folder):
            exts = {'.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif'}
            if not os.path.isdir(folder):
                return []
            return [f for f in os.listdir(folder)
                    if os.path.splitext(f)[1].lower() in exts]

        defect_imgs     = _list_images(defects_path)
        non_defect_imgs = _list_images(non_defects_path)
        all_imgs = ([(f, 'defects') for f in defect_imgs] +
                    [(f, 'non_defects') for f in non_defect_imgs])
        all_imgs.sort(key=lambda x: x[0])

        HDR_FILL = PatternFill('solid', fgColor='2D7DD2')
        HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
        DEF_FILL = PatternFill('solid', fgColor='F8D7DA')
        NOD_FILL = PatternFill('solid', fgColor='D4EDDA')
        THIN     = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
        CENTER   = Alignment(horizontal='center', vertical='center')
        LEFT     = Alignment(horizontal='left', vertical='center')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Batch Results'

        CAT_LABELS = {'defects': 'Indicator', 'non_defects': 'Non Indicator'}

        headers = ['Image Name', 'Classification', 'Number of Segments']
        widths  = [36, 20, 20]
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = CENTER; c.border = THIN
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        for r_idx, (fname, cat) in enumerate(all_imgs, start=2):
            fill = DEF_FILL if cat == 'defects' else NOD_FILL
            num_segments = segment_counts.get(fname, 0)
            label = CAT_LABELS.get(cat, cat)
            for col, val in enumerate([fname, label, num_segments], start=1):
                c = ws.cell(row=r_idx, column=col, value=val)
                c.fill = fill; c.border = THIN
                c.alignment = LEFT if col == 1 else CENTER

        ws2 = wb.create_sheet(title='Summary')
        summary_data = [
            ('Total Images',   len(all_imgs)),
            ('Indicators',     len(defect_imgs)),
            ('Non Indicators', len(non_defect_imgs)),
        ]
        for r, (k, v) in enumerate(summary_data, start=1):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=v)

        wb.save(report_path)

    def _batch3d_run_viz(self, output_base, stl_or_step_path, auto=False, seg_nrrd_path=None):
        """Launch 3D Slicer visualization (mirrors original _launch_3d logic).

        Change 7: when `seg_nrrd_path` is provided (the new "seg.nrrd + STL"
        workflow), the existing headless segmentation-GENERATION phase is
        skipped entirely — that phase's only job is to produce a seg.nrrd
        file from the PNG stack, and the caller is already handing us one.
        We reuse the user-supplied seg.nrrd in place of the generated one
        and jump straight to the SAME visualization phase (`visualise()`
        in the Slicer script below) that the Results-Folder workflow uses:
        identical transform constants, identical Harden Transform call,
        identical final view. No second visualization pipeline is created.
        """
        import glob as _glob, subprocess, tempfile

        slicer_exe = self._slicer_exe_var.get().strip()
        if not slicer_exe or not os.path.isfile(slicer_exe):
            msg = (f'3D Slicer executable not found:\n  {slicer_exe}\n\n'
                   'Place Slicer inside <project>/slicer/ folder.')
            self.root.after(0, lambda: messagebox.showwarning('3D Slicer Not Found', msg))
            self.root.after(0, lambda: self._batch3d_set_progress(100))
            self.root.after(0, lambda: self._batch3d_status_lbl.config(text='3D Viz skipped (Slicer not found)'))
            return

        # Results folder = output_base (contains defects / non_defects)
        folder = output_base

        # Determine STL path (STEP → STL conversion if needed)
        step_path = Path(stl_or_step_path)
        if step_path.suffix.lower() in ('.step', '.stp'):
            stl_file = str(step_path.with_suffix('.stl'))
            step_file = str(step_path)
        else:
            stl_file  = str(step_path)
            step_file = str(step_path)

        seg_out = os.path.join(output_base, '3D_Models', 'indicator_model.seg.nrrd')
        os.makedirs(os.path.join(output_base, '3D_Models'), exist_ok=True)

        # ── Change 7: direct seg.nrrd + STL workflow ───────────────────────
        # If the caller already has a seg.nrrd on disk (new workflow), skip
        # PNG collection / headless generation altogether and go straight
        # to the visualization-only Slicer pass, reusing it as-is.
        skip_generation = bool(seg_nrrd_path)
        if skip_generation:
            seg_src = str(seg_nrrd_path)
            if not os.path.isfile(seg_src):
                self.root.after(0, lambda: messagebox.showerror(
                    '3D Visualization Error', f'seg.nrrd file not found:\n{seg_src}'))
                self.root.after(0, lambda: self._batch3d_set_progress(100))
                self.root.after(0, lambda: self._batch3d_status_lbl.config(text='3D Viz failed (seg.nrrd missing)'))
                return
            # Use the supplied seg.nrrd directly as the file the Slicer
            # visualization pass loads (seg_output_path below). If it isn't
            # already inside output_base/3D_Models, copy it there so the
            # rest of the pipeline (and the STL copy step) behaves exactly
            # like the existing Results-Folder workflow.
            try:
                if os.path.abspath(seg_src) != os.path.abspath(seg_out):
                    import shutil as _shutil2
                    _shutil2.copy2(seg_src, seg_out)
                self.root.after(0, lambda p=seg_out: self.tab3_log(f"seg.nrrd ready for visualization → {p}"))
            except Exception as _seg_copy_exc:
                self.root.after(0, lambda e=_seg_copy_exc: self.tab3_log(
                    f"⚠ Could not stage seg.nrrd into 3D Models folder: {e}"))
                seg_out = seg_src  # fall back to loading it directly in place

            png_files = []
            category_summary = ''
        else:
            # Collect PNGs from defects / non_defects subfolders
            png_entries = []
            for sub in sorted(os.listdir(folder)):
                sub_path = os.path.join(folder, sub)
                if os.path.isdir(sub_path) and sub.lower() in ('defects', 'non_defects', 'defect', 'non_defect'):
                    for fname in sorted(os.listdir(sub_path)):
                        if fname.lower().endswith('.png'):
                            png_entries.append((sub, os.path.join(sub_path, fname)))

            if not png_entries:
                self.root.after(0, lambda: self.tab3_log(
                    "⚠ No PNG files found for 3D Visualization."))
                self.root.after(0, lambda: self._batch3d_set_progress(100))
                self.root.after(0, lambda: self._batch3d_status_lbl.config(text='3D Viz skipped (no PNGs)'))
                return

            png_files = [path for _, path in png_entries]
            from collections import Counter as _Counter
            cat_counts = _Counter(cat for cat, _ in png_entries)
            category_summary = '\n'.join(
                f'    {cat}: {cnt} image(s)' for cat, cnt in sorted(cat_counts.items()))

        # FreeCAD conversion (same as original)
        freecad_exe = self._freecad_exe_var.get().strip()
        if freecad_exe and os.path.isfile(freecad_exe) and step_file != stl_file:
            fc_script_text = f"""
import FreeCAD, Part, Mesh, MeshPart
doc = FreeCAD.newDocument()
shape = Part.read(r"{step_file}")
obj = doc.addObject("Part::Feature", "Part")
obj.Shape = shape
mesh = MeshPart.meshFromShape(Shape=shape, LinearDeflection=0.1, AngularDeflection=0.5, Relative=False)
mesh_obj = doc.addObject("Mesh::Feature", "Mesh")
mesh_obj.Mesh = mesh
Mesh.export([mesh_obj], r"{stl_file}")
"""
            fc_tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False,
                prefix='fc_convert_', encoding='utf-8')
            fc_tmp.write(fc_script_text); fc_tmp.close()
            try:
                subprocess.run([freecad_exe, fc_tmp.name], timeout=120, capture_output=True)
            except Exception:
                pass

        # ── Change 6: copy the STL used for visualization into 3D_Models
        #    too, so seg.nrrd and the STL always exist together. ───────────
        try:
            import shutil as _shutil
            if os.path.isfile(stl_file):
                _stl_dest = os.path.join(output_base, '3D_Models', os.path.basename(stl_file))
                if os.path.abspath(stl_file) != os.path.abspath(_stl_dest):
                    _shutil.copy2(stl_file, _stl_dest)
                self.root.after(0, lambda p=_stl_dest: self.tab3_log(f"STL copied to 3D Models folder → {p}"))
            else:
                self.root.after(0, lambda: self.tab3_log("⚠ STL file not found yet — skipping copy to 3D Models folder."))
        except Exception as _copy_exc:
            self.root.after(0, lambda e=_copy_exc: self.tab3_log(f"⚠ Could not copy STL into 3D Models folder: {e}"))

        # Build Slicer script (identical to original _launch_3d, now
        # parameterised by `headless` to split generation from display —
        # see Change 2: Slicer Launch Timing).
        import textwrap
        sp_x=self._VIZ3D_SP_X; sp_y=self._VIZ3D_SP_Y; sp_z=self._VIZ3D_SP_Z
        island=self._VIZ3D_ISLAND; r_min=self._VIZ3D_R_MIN; g_max=self._VIZ3D_G_MAX
        sx=self._VIZ3D_SX; sy=self._VIZ3D_SY; sz=self._VIZ3D_SZ
        tx=self._VIZ3D_TX; ty=self._VIZ3D_TY; tz=self._VIZ3D_TZ

        import textwrap as _textwrap

        def _make_script(headless):
            return _textwrap.dedent(f'''
            import os, glob, subprocess, random, pathlib, slicer, numpy as np, qt, vtk

            try:
                print(f"VTK thread pool size: {{vtk.vtkSlicerThreading.GetNumberOfThreads()}}")
            except Exception:
                pass
            os.environ['VTK_NUM_THREADS'] = '32'
            try:
                slicer.app.threading.setNumberOfThreads(32)
            except Exception:
                pass
            
            png_folder = r"{folder}"
            seg_output_path = r"{seg_out}"
            png_file_list = {repr(png_files)}
            spacing_x = {sp_x}; spacing_y = {sp_y}; spacing_z = {sp_z}
            minimum_island_size = {island}
            R_min, G_max, B_max = {r_min}, {g_max}, 0
            step_file = pathlib.Path(r"{step_file}")
            stl_file  = pathlib.Path(r"{stl_file}")
            freecad_cmd = r"{freecad_exe if freecad_exe else ''}"

            def create_segmentation_from_png():
                print("STAGE::LOADING_DATA")
                print("STEP::1/13::Loading PNG Stack...")
                print("\\n=== Loading PNG stack ===")
                png_files = png_file_list if png_file_list else sorted(
                    glob.glob(os.path.join(png_folder, "**", "*.png"), recursive=True))
                if not png_files:
                    raise RuntimeError(f"No PNG files found in {{png_folder}}")
                print(f"Total PNGs: {{len(png_files)}}")
                setup_view()
                from PIL import Image
                slices = []; target_size = None
                for f in png_files:
                    img = Image.open(f).convert("RGB")
                    if target_size is None: target_size = img.size
                    if img.size != target_size: img = img.resize(target_size)
                    slices.append(np.array(img))
                if not slices: raise RuntimeError("No valid PNG images loaded")
                volume_array = np.stack(slices, axis=0)
                volume_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLVectorVolumeNode", "PNGStack")
                slicer.util.updateVolumeFromArray(volume_node, volume_array)
                volume_node.SetSpacing(spacing_x, spacing_y, spacing_z)
                volume_node.CreateDefaultDisplayNodes()
                # ── Orientation correction: flip X axis so IJK→RAS becomes IJK→LAS ──
                _vtk_mat = vtk.vtkMatrix4x4()
                volume_node.GetIJKToRASMatrix(_vtk_mat)
                _current = np.array(
                    [[_vtk_mat.GetElement(i, j) for j in range(4)] for i in range(4)],
                    dtype=np.float64)
                _flip_x = np.diag([-1.0, 1.0, 1.0, 1.0])
                _new_mat_np = _current @ _flip_x
                _new_vtk_mat = vtk.vtkMatrix4x4()
                for i in range(4):
                    for j in range(4):
                        _new_vtk_mat.SetElement(i, j, float(_new_mat_np[i, j]))
                volume_node.SetIJKToRASMatrix(_new_vtk_mat)
                volume_node.Modified()
                print("IJK->LAS orientation correction applied to PNGStack.")
                # ── End orientation correction ─────────────────────────────────────
                rgb_arr = slicer.util.arrayFromVolume(volume_node)
                print("STEP::2/13::Extracting Red Mask...")
                R = rgb_arr[..., 0] if rgb_arr.shape[-1]==3 else rgb_arr[0]
                G = rgb_arr[..., 1] if rgb_arr.shape[-1]==3 else rgb_arr[1]
                B = rgb_arr[..., 2] if rgb_arr.shape[-1]==3 else rgb_arr[2]
                mask = ((R==R_min)&(G==G_max)&(B==B_max)).astype(np.uint8)
                print("STEP::3/13::Creating Scalar Volume...")
                scalar_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLScalarVolumeNode", "RedMaskVolume")
                slicer.util.updateVolumeFromArray(scalar_node, mask)
                scalar_node.CopyOrientation(volume_node); scalar_node.SetOrigin(volume_node.GetOrigin())
                scalar_node.SetSpacing(volume_node.GetSpacing()); scalar_node.CreateDefaultDisplayNodes()
                print("STAGE::SEGMENTATION")
                print("STEP::4/13::Creating Segmentation...")
                seg_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode", "IndicatorSegmentation")
                seg_node.CreateDefaultDisplayNodes()
                seg_node.SetReferenceImageGeometryParameterFromVolumeNode(scalar_node)
                segmentation = seg_node.GetSegmentation()
                segment_id = segmentation.AddEmptySegment("Indicator")
                seg_node.CreateBinaryLabelmapRepresentation()
                ew = slicer.qMRMLSegmentEditorWidget(); ew.setMRMLScene(slicer.mrmlScene)
                en = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentEditorNode")
                ew.setMRMLSegmentEditorNode(en); ew.setSegmentationNode(seg_node); ew.setSourceVolumeNode(scalar_node)
                en.SetSelectedSegmentID(segment_id)
                ew.setActiveEffectByName("Threshold"); t = ew.activeEffect()
                print("STEP::5/13::Applying Threshold...")
                t.setParameter("MinimumThreshold","1"); t.setParameter("MaximumThreshold","255"); t.self().onApply()
                slicer.app.processEvents()
                ew.setActiveEffectByName("Islands"); isl = ew.activeEffect()
                print("STEP::6/13::Removing Small Islands...")
                isl.setParameter("Operation","REMOVE_SMALL_ISLANDS"); isl.setParameter("MinimumSize",str(minimum_island_size))
                isl.self().onApply(); slicer.app.processEvents()
                for i in range(segmentation.GetNumberOfSegments()):
                    seg_id = segmentation.GetNthSegmentID(i)
                    segmentation.GetSegment(seg_id).SetColor(random.random(),random.random(),random.random())
                print("STAGE::MODEL_GENERATION")
                print("STEP::7/13::Generating Closed Surface...")
                seg_node.CreateClosedSurfaceRepresentation(); seg_node.GetDisplayNode().SetVisibility3D(True)
                print("STAGE::STL_PROCESSING")
                print("STEP::8/13::Saving Segmentation (.seg.nrrd)...")
                slicer.util.saveNode(seg_node, seg_output_path)
                print(f"Segmentation saved to {{seg_output_path}}")
                return seg_node

            def visualise(seg_node, stl_path):
                print("STAGE::VISUALIZATION")
                print("STEP::9/13::Loading Segmentation...")
                segNode = slicer.util.loadSegmentation(seg_output_path); segNode.CreateClosedSurfaceRepresentation()
                sd = segNode.GetDisplayNode(); sd.SetVisibility3D(True); sd.SetOpacity3D(1)
                print("STEP::10/13::Loading STL Model...")
                modelNode = slicer.util.loadModel(stl_path); modelNode.CreateDefaultDisplayNodes()
                md = modelNode.GetDisplayNode(); md.SetColor(1,0,0); md.SetOpacity(0.1)
                tn = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLinearTransformNode","ScaleTranslate")
                print("STEP::11/13::Scaling & Translating STL...")
                mx = vtk.vtkMatrix4x4(); mx.Identity()
                mx.SetElement(0,0,{sx}); mx.SetElement(1,1,{sy}); mx.SetElement(2,2,{sz})
                mx.SetElement(0,3,{tx}); mx.SetElement(1,3,{ty}); mx.SetElement(2,3,{tz})
                tn.SetMatrixTransformToParent(mx); modelNode.SetAndObserveTransformNodeID(tn.GetID())
                print("STEP::12/13::Hardening Transform...")
                slicer.vtkSlicerTransformLogic().hardenTransform(modelNode)
                print("STEP::13/13::Launching 3D Visualization...")
                slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUp3DView)
                v = slicer.app.layoutManager().threeDWidget(0).threeDView(); v.resetFocalPoint(); v.resetCamera()
                print("Visualization ready")

            def setup_view():
                if HEADLESS:
                    # No main window exists under --no-main-window; skip
                    # window-chrome setup entirely during the headless pass.
                    return
                mw = slicer.util.mainWindow()
                slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutOneUp3DView)
                for d in mw.findChildren("QDockWidget"): d.hide()
                mw.menuBar().hide()
                for t in mw.findChildren("QToolBar"): t.hide()
                mw.statusBar().hide(); mw.showMaximized()

            HEADLESS = {('True' if headless else 'False')}

            def delayed_run():
                try:
                    if HEADLESS:
                        # Generation-only pass: build & save the segmentation
                        # (.seg.nrrd) WITHOUT showing/using the 3D view.
                        # The Slicer window is never displayed to the user
                        # during this phase — it only opens afterwards, once
                        # generation has finished, for the visualization pass.
                        create_segmentation_from_png()
                        print("STAGE::COMPLETE")
                        qt.QTimer.singleShot(800, lambda: slicer.app.exit(0))
                    else:
                        # Visualization-only pass: seg.nrrd already exists on
                        # disk from the prior headless generation phase.
                        setup_view()
                        visualise(None, stl_file)
                        print("STAGE::COMPLETE")
                except Exception as _e:
                    print(f"STAGE::ERROR::{{_e}}")
                    if HEADLESS:
                        slicer.app.exit(1)
                    raise

            qt.QTimer.singleShot(200, delayed_run)
        ''')

        import tempfile as _tempfile

        try:
            import subprocess as _subprocess
            env = os.environ.copy()
            for _k in ('QT_QPA_PLATFORM_PLUGIN_PATH', 'QT_PLUGIN_PATH', 'QT_DEBUG_PLUGINS'):
                env.pop(_k, None)
            _slicer_root = os.path.dirname(os.path.dirname(os.path.abspath(slicer_exe)))
            for _cand in (os.path.join(_slicer_root, 'lib', 'Qt', 'plugins'),
                          os.path.join(_slicer_root, 'lib', 'qt5', 'plugins'),
                          os.path.join(_slicer_root, 'plugins')):
                if os.path.isdir(_cand):
                    env['QT_QPA_PLATFORM_PLUGIN_PATH'] = _cand; break
            env['QT_QPA_PLATFORM'] = 'xcb'; env['XDG_SESSION_TYPE'] = 'x11'

            def _set_progress(p):
                self._batch3d_set_progress(p)

            def _set_status(txt):
                self._batch3d_status_lbl.config(text=txt)

            # Progress range for the 3D Viz phase itself spans 75→100 when
            # launched as part of the auto chain (matches existing phase
            # hints), or starts fresh at 0→100 when launched independently.
            # The phase is further split between the headless generation
            # pass (Change 2) and the GUI visualization pass so the bar
            # keeps advancing smoothly with no single big jump.
            base, span = (75, 25) if auto else (0, 100)
            gen_base, gen_span = base, span * 0.6
            viz_base, viz_span = base + span * 0.6, span * 0.4

            gen_stage_map = {
                'LOADING_DATA':     (gen_base + gen_span * 0.20, 'Loading data…'),
                'SEGMENTATION':     (gen_base + gen_span * 0.55, 'Segmentation…'),
                'MODEL_GENERATION': (gen_base + gen_span * 0.80, 'Model generation…'),
                'STL_PROCESSING':   (gen_base + gen_span * 0.95, 'Saving segmentation…'),
            }
            viz_stage_map = {
                'VISUALIZATION':    (viz_base + viz_span * 0.6, 'Visualization…'),
            }

            def _launch_phase(headless, stage_map, on_done):
                """Launch one Slicer subprocess (generation or visualization
                pass), stream its output, and invoke on_done(ok, error) on
                the GUI thread once it exits. When headless=True, Slicer
                runs with --no-main-window so its window is never shown —
                it only opens for the visualization pass, and only after
                generation has completed successfully (Change 2)."""
                script_content = _make_script(headless)
                tmp = _tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False,
                    prefix='indicator_nrrd_stl_', encoding='utf-8')
                tmp.write(script_content); tmp.close()

                cmd = [slicer_exe, '--no-splash']
                if headless:
                    cmd.append('--no-main-window')
                cmd += ['--python-script', tmp.name]

                proc = _subprocess.Popen(
                    cmd, shell=False, env=env,
                    stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT,
                    text=True, bufsize=1)

                def _monitor():
                    completed_ok = False
                    error_seen   = None
                    try:
                        for line in proc.stdout:
                            line = line.rstrip('\n')
                            if not line:
                                continue
                            if line.startswith('STAGE::ERROR::'):
                                error_seen = line[len('STAGE::ERROR::'):]
                                self.root.after(0, lambda m=line: self.tab3_log(f"⚠ {m}"))
                                continue
                            if line.startswith('STAGE::COMPLETE'):
                                completed_ok = True
                                self.root.after(0, lambda: self.tab3_log("Pipeline stage reported completion."))
                                continue
                            if line.startswith('STEP::'):
                                step_body = line[len('STEP::'):]
                                step_no, _, step_label = step_body.partition('::')
                                display = f"[{step_no}] {step_label}" if step_label else f"[{step_no}]"
                                self.root.after(0, lambda d=display: self.tab3_log(d))
                                continue
                            if line.startswith('STAGE::'):
                                stage_key = line[len('STAGE::'):]
                                pct, label = stage_map.get(stage_key, (None, None))
                                if pct is not None:
                                    self.root.after(0, lambda p=pct: _set_progress(p))
                                    self.root.after(0, lambda l=label: _set_status(l))
                                    self.root.after(0, lambda l=label: self.tab3_log(f"… {l}"))
                                continue
                            self.root.after(0, lambda m=line: self.tab3_log(f"[Slicer] {m}"))
                    except Exception as _read_exc:
                        self.root.after(0, lambda e=_read_exc: self.tab3_log(f"⚠ Lost connection to Slicer output stream: {e}"))

                    # ── Capture the real process exit code so a crash (e.g. a
                    #    native segfault/abort with no Python-catchable
                    #    exception) is visible instead of a silent
                    #    "unknown error". On POSIX, a negative return code
                    #    (-N) means the process was killed by signal N
                    #    (e.g. -11 = SIGSEGV, -6 = SIGABRT). ───────────────
                    try:
                        rc = proc.wait(timeout=5)
                    except Exception:
                        rc = proc.poll()
                    if rc is not None and rc != 0:
                        if rc < 0:
                            self.root.after(0, lambda c=rc: self.tab3_log(
                                f"⚠ Slicer process was terminated by signal {-c} (exit code {c}) — likely a native crash."))
                        else:
                            self.root.after(0, lambda c=rc: self.tab3_log(
                                f"⚠ Slicer process exited with non-zero code {c}."))
                        if not error_seen:
                            error_seen = f"Slicer exited abnormally (code {rc})"

                    self.root.after(0, lambda: on_done(completed_ok, error_seen))

                threading.Thread(target=_monitor, daemon=True).start()

            def _on_viz_done(ok, err):
                if ok and not err:
                    _set_progress(100)
                    _set_status('3D Visualization complete!')
                    self.tab3_log("✓ 3D Visualization finished successfully — Slicer window is ready for interaction.")
                elif err:
                    _set_status('3D Visualization failed.')
                    self.tab3_log(f"✗ 3D Visualization failed: {err}")
                    messagebox.showerror('3D Visualization Error',
                                          f'The visualization pipeline reported an error:\n{err}')
                else:
                    _set_status('3D Visualization ended unexpectedly.')
                    self.tab3_log("⚠ Slicer output stream closed before the pipeline reported completion. "
                                  "3D Visualization may not have finished successfully.")

            def _on_gen_done(ok, err):
                if not ok or err:
                    _set_status('3D segmentation generation failed.')
                    self.tab3_log(f"✗ Segmentation generation failed: {err or 'unknown error'} — Slicer will not be opened.")
                    messagebox.showerror('3D Visualization Error',
                                          f'Segmentation generation failed:\n{err or "unknown error"}\n\n'
                                          f'3D Slicer was not opened.')
                    return
                # Required files (seg.nrrd + STL) are now confirmed on disk —
                # only now does the Slicer window open, for visualization.
                self.tab3_log("✓ Segmentation generated and saved — required files are ready.")
                self.tab3_log("Opening 3D Slicer for visualization…")
                _set_status('Opening 3D Slicer…')
                _set_progress(viz_base)
                _launch_phase(False, viz_stage_map, _on_viz_done)

            if skip_generation:
                # Change 7: seg.nrrd already supplied by the user — there is
                # nothing to generate, so the headless pass is skipped and
                # we go straight to the SAME visualization phase used by
                # the Results-Folder workflow.
                self.root.after(0, lambda: _set_status('Opening 3D Slicer…'))
                self.root.after(0, lambda: self.tab3_log(
                    "\n3D Visualization pipeline started — using the supplied seg.nrrd directly "
                    "(no segmentation generation needed)…"))
                self.root.after(0, lambda: _set_progress(viz_base))
                _launch_phase(False, viz_stage_map, _on_viz_done)
            else:
                self.root.after(0, lambda: _set_status('Generating 3D segmentation…'))
                self.root.after(0, lambda: self.tab3_log(
                    "\n3D Visualization pipeline started — generating segmentation in the background "
                    "(Slicer window will open once generation finishes)…"))
                self.root.after(0, lambda: _set_progress(gen_base))

                _launch_phase(True, gen_stage_map, _on_gen_done)

        except Exception as exc:
            def _viz_err(e=exc):
                messagebox.showerror('Launch Error', f'Failed to launch 3D Slicer:\n{e}')
                self._batch3d_set_progress(100)
                self._batch3d_status_lbl.config(text='3D Viz launch failed.')
                self.tab3_log(f"✗ Failed to launch 3D Slicer: {e}")
            self.root.after(0, _viz_err)

    def _batch3d_launch_viz_independent(self):
        """Independent 3D Visualise button — opens a dedicated dialog to pick
        inputs for visualization, completely independent of the main batch
        inputs.

        Change 7: the dialog now offers two input modes that both feed the
        SAME underlying `_batch3d_run_viz` visualization pipeline:
          • Results Folder + STL  (original workflow — unchanged)
          • seg.nrrd + STL        (new workflow — produces an identical
                                    visualization, per the existing transform
                                    + Harden Transform logic already in
                                    `_batch3d_run_viz`)
        No second visualization implementation is created; only the input
        mechanism differs.
        """

        # ── Modal dialog ─────────────────────────────────────────────────
        popup = tk.Toplevel(self.root)
        popup.title("3D Visualise — Select Inputs")
        popup.resizable(False, False)
        popup.configure(bg=self.BG)
        popup.grab_set()

        # Centre over main window
        self.root.update_idletasks()
        DIALOG_W, DIALOG_H = 540, 360
        rx = self.root.winfo_x() + (self.root.winfo_width()  - DIALOG_W) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - DIALOG_H) // 2
        popup.geometry(f"{DIALOG_W}x{DIALOG_H}+{rx}+{ry}")

        card = tk.Frame(popup, bg=self.PANEL)
        card.place(relx=0.5, rely=0.5, anchor='center', width=500, height=330)

        tk.Label(card, text='3D Visualise', bg=self.PANEL, fg=self.HIGHLIGHT,
                 font=('Segoe UI', 13, 'bold')).pack(anchor='w', padx=20, pady=(16, 2))
        tk.Label(card, text='Choose an input mode, then select files to launch visualization independently.',
                 bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), wraplength=460, justify='left').pack(anchor='w', padx=20, pady=(0, 8))

        # ── Mode toggle: Results Folder + STL  /  seg.nrrd + STL ──────────
        mode_var = tk.StringVar(value='seg_nrrd')   # default to the new workflow

        mode_row = tk.Frame(card, bg=self.PANEL)
        mode_row.pack(fill='x', padx=20, pady=(0, 8))
        tk.Radiobutton(mode_row, text='seg.nrrd + STL', variable=mode_var, value='seg_nrrd',
                       bg=self.PANEL, fg=self.FG, selectcolor=self.ENTRY_BG,
                       activebackground=self.PANEL, activeforeground=self.HIGHLIGHT,
                       font=('Segoe UI', 9, 'bold'), cursor='hand2',
                       command=lambda: _set_mode('seg_nrrd')).pack(side='left', padx=(0, 16))
        tk.Radiobutton(mode_row, text='Results Folder + STL', variable=mode_var, value='results',
                       bg=self.PANEL, fg=self.FG, selectcolor=self.ENTRY_BG,
                       activebackground=self.PANEL, activeforeground=self.HIGHLIGHT,
                       font=('Segoe UI', 9, 'bold'), cursor='hand2',
                       command=lambda: _set_mode('results')).pack(side='left')

        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=20, pady=(0, 10))

        # ── Results Folder row (original workflow) ─────────────────────────
        results_var = tk.StringVar(value='No folder selected')

        def _pick_results():
            p = filedialog.askdirectory(
                title='Select Results Folder  (must contain Indicator / Non Indicator subfolders)',
                parent=popup)
            if p:
                results_var.set(p)

        row1 = tk.Frame(card, bg=self.PANEL)
        tk.Label(row1, text='Results Folder :', bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), width=16, anchor='w').pack(side='left')
        tk.Entry(row1, textvariable=results_var, bg=self.ENTRY_BG, fg=self.FG,
                 relief='flat', font=('Segoe UI', 9), state='readonly',
                 readonlybackground=self.ENTRY_BG, bd=0
                 ).pack(side='left', fill='x', expand=True, ipady=5, padx=(4, 8))
        tk.Button(row1, text='Browse', command=_pick_results,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=10,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

        # ── seg.nrrd row (new workflow — Change 7) ──────────────────────────
        seg_var = tk.StringVar(value='No file selected')

        def _pick_seg():
            p = filedialog.askopenfilename(
                title='Select seg.nrrd File',
                filetypes=[('Segmentation NRRD files', '*.nrrd'), ('All files', '*.*')],
                parent=popup)
            if p:
                seg_var.set(p)

        row_seg = tk.Frame(card, bg=self.PANEL)
        tk.Label(row_seg, text='seg.nrrd File :', bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), width=16, anchor='w').pack(side='left')
        tk.Entry(row_seg, textvariable=seg_var, bg=self.ENTRY_BG, fg=self.FG,
                 relief='flat', font=('Segoe UI', 9), state='readonly',
                 readonlybackground=self.ENTRY_BG, bd=0
                 ).pack(side='left', fill='x', expand=True, ipady=5, padx=(4, 8))
        tk.Button(row_seg, text='Browse', command=_pick_seg,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=10,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

        # ── STL File row (shared by both modes) ─────────────────────────────
        stl_var = tk.StringVar(value='No file selected')

        def _pick_stl():
            p = filedialog.askopenfilename(
                title='Select STL / STEP File',
                filetypes=[('STL/STEP files', '*.stl *.step *.stp'), ('All files', '*.*')],
                parent=popup)
            if p:
                stl_var.set(p)

        row2 = tk.Frame(card, bg=self.PANEL)
        row2.pack(fill='x', padx=20, pady=4)
        tk.Label(row2, text='STL / STEP File :', bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), width=16, anchor='w').pack(side='left')
        tk.Entry(row2, textvariable=stl_var, bg=self.ENTRY_BG, fg=self.FG,
                 relief='flat', font=('Segoe UI', 9), state='readonly',
                 readonlybackground=self.ENTRY_BG, bd=0
                 ).pack(side='left', fill='x', expand=True, ipady=5, padx=(4, 8))
        tk.Button(row2, text='Browse', command=_pick_stl,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=10,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

        def _set_mode(mode):
            mode_var.set(mode)
            if mode == 'seg_nrrd':
                row1.pack_forget()
                row_seg.pack(fill='x', padx=20, pady=4, before=row2)
            else:
                row_seg.pack_forget()
                row1.pack(fill='x', padx=20, pady=4, before=row2)
            err_lbl.config(text='')

        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=20, pady=(10, 8))

        err_lbl = tk.Label(card, text='', bg=self.PANEL, fg='#e05c5c',
                           font=('Segoe UI', 9), wraplength=460, justify='left')
        err_lbl.pack(pady=(0, 4))

        # Show the default mode's row now that err_lbl/row2 exist
        _set_mode('seg_nrrd')

        # ── Launch button ─────────────────────────────────────────────────
        def _launch():
            mode     = mode_var.get()
            stl_path = stl_var.get().strip()

            if stl_path == 'No file selected' or not os.path.isfile(stl_path):
                err_lbl.config(text='Please select a valid STL / STEP File.')
                return

            if mode == 'seg_nrrd':
                seg_path = seg_var.get().strip()
                if seg_path == 'No file selected' or not os.path.isfile(seg_path):
                    err_lbl.config(text='Please select a valid seg.nrrd File.')
                    return
                if not seg_path.lower().endswith('.nrrd'):
                    err_lbl.config(text='Selected file does not look like a .nrrd file.')
                    return

                popup.destroy()

                self.tab3_log(f"\n3D Visualization started (seg.nrrd + STL mode)")
                self.tab3_log(f"  seg.nrrd : {seg_path}")
                self.tab3_log(f"  STL      : {stl_path}")
                self._batch3d_status_lbl.config(text='3D Visualization…')
                self._batch3d_set_progress(0)

                # Use the STL's own folder as the output_base, so the
                # staged copy of seg.nrrd + STL lands in
                # <stl folder>/3D_Models — same convention the existing
                # Results-Folder workflow already uses (output_base/3D_Models).
                output_base = os.path.dirname(os.path.abspath(stl_path))

                def _run_viz():
                    self._batch3d_run_viz(output_base, stl_path, auto=False,
                                           seg_nrrd_path=seg_path)
                threading.Thread(target=_run_viz, daemon=True).start()

            else:
                results_path = results_var.get().strip()
                if results_path == 'No folder selected' or not os.path.isdir(results_path):
                    err_lbl.config(text='Please select a valid Results Folder.')
                    return

                # Verify at least one of defects / non_defects exists
                has_sub = any(
                    os.path.isdir(os.path.join(results_path, s))
                    for s in ('defects', 'non_defects', 'Defects', 'Non_Defects')
                )
                if not has_sub:
                    err_lbl.config(
                        text='Results Folder must contain Indicator / Non Indicator subfolders.')
                    return

                popup.destroy()

                self.tab3_log(f"\n3D Visualization started (Results Folder mode)")
                self.tab3_log(f"  Results : {results_path}")
                self.tab3_log(f"  STL     : {stl_path}")
                self._batch3d_status_lbl.config(text='3D Visualization…')
                self._batch3d_set_progress(0)

                def _run_viz():
                    self._batch3d_run_viz(results_path, stl_path, auto=False)
                threading.Thread(target=_run_viz, daemon=True).start()

        btn_row = tk.Frame(card, bg=self.PANEL)
        btn_row.pack(pady=(4, 0))
        tk.Button(btn_row, text='  Launch 3D Viewer',
                  command=_launch,
                  bg='#1e3a2f', fg='#5dbf8e', relief='flat',
                  font=('Segoe UI', 11, 'bold'), padx=22, pady=7,
                  activebackground='#14532d', cursor='hand2').pack(side='left', padx=(0, 10))
        tk.Button(btn_row, text='Cancel',
                  command=popup.destroy,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 10, 'bold'), padx=14, pady=7,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='left')

    # ==================== TAB 1: ROI SELECTION (popup window) ====================

    def open_roi_window(self):
        """Open ROI Selection as a top-level window."""
        if self._roi_win is not None and self._roi_win.winfo_exists():
            self._roi_win.lift()
            return

        win = tk.Toplevel(self.root)
        win.title("ROI Selection")
        win.geometry("900x650")
        win.configure(bg=self.BG)
        self._roi_win = win

        ctrl = tk.Frame(win, bg=self.PANEL, height=44)
        ctrl.pack(side='top', fill='x')
        ctrl.pack_propagate(False)

        for txt, cmd in [("Upload Image",     self.tab1_upload_image),
                         ("Clear Points",     self.tab1_clear_points),
                         ("Save Coordinates", self.tab1_save_coordinates)]:
            tk.Button(ctrl, text=txt, command=cmd,
                      bg=self.BTN_BG, fg=self.FG, relief='flat',
                      font=('Segoe UI', 9, 'bold'), padx=10, pady=4,
                      activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='left', padx=6, pady=6)

        self.tab1_info_label = tk.Label(ctrl,
            text="Click on image to select polygon points (click on first point to close)",
            bg=self.PANEL, fg=self.FG2, font=('Segoe UI', 9))
        self.tab1_info_label.pack(side='left', padx=16)

        canvas_frame = tk.Frame(win, bg=self.BG)
        canvas_frame.pack(fill='both', expand=True, padx=5, pady=5)

        self.tab1_canvas = tk.Canvas(canvas_frame, bg='#181d22', highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(canvas_frame, orient='vertical',
                                    command=self.tab1_canvas.yview)
        h_scrollbar = ttk.Scrollbar(canvas_frame, orient='horizontal',
                                    command=self.tab1_canvas.xview)
        self.tab1_canvas.configure(yscrollcommand=v_scrollbar.set,
                                   xscrollcommand=h_scrollbar.set)
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')
        self.tab1_canvas.pack(side='left', fill='both', expand=True)

        self.tab1_canvas.bind('<Button-1>', self.tab1_canvas_click)
        self.tab1_canvas.bind('<MouseWheel>',
                              lambda e: self.tab1_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.tab1_canvas.bind('<Shift-MouseWheel>',
                              lambda e: self.tab1_canvas.xview_scroll(int(-1*(e.delta/120)), "units"))
        self.tab1_canvas.bind('<Control-MouseWheel>', self.tab1_zoom)
        self.tab1_canvas.bind('<plus>',  lambda e: self.tab1_zoom_button(1.1))
        self.tab1_canvas.bind('<minus>', lambda e: self.tab1_zoom_button(0.9))
        self.tab1_canvas.bind('<equal>', lambda e: self.tab1_zoom_button(1.1))

        if self.tab1_image is not None:
            self.tab1_display_image()

    def init_tab1(self):
        """Initialise ROI Selection variables (UI is opened via open_roi_window)."""
        self.tab1_image = None
        self.tab1_cv_image = None
        self.tab1_photo = None
        self.tab1_polygon_points = []
        self.tab1_image_path = None
        self.tab1_polygon_closed = False
        self.tab1_zoom_level = 1.0
        self._roi_win = None
        # Placeholder label – real UI lives in open_roi_window()
        self.tab1_info_label = tk.Label(self.root)  # dummy; replaced when window opens
        

    def tab1_upload_image(self):
        """Upload image for ROI selection"""
        file_path = filedialog.askopenfilename(
            title="Select Image",
            filetypes=[("Image files", "*.jpg *.jpeg *.png"),
                       ("All files", "*.*")]
        )

        if file_path:
            self.tab1_image_path = file_path
            self.tab1_cv_image = cv2.imread(file_path)
            self.tab1_image = cv2.cvtColor(self.tab1_cv_image, cv2.COLOR_BGR2RGB)
            self.tab1_polygon_points = []

            self.tab1_polygon_closed = False
            self.tab1_display_image()

    
    def tab1_zoom(self, event):
        """Zoom image with Ctrl+MouseWheel"""
        if self.tab1_image is None:
            return
        
        factor = 1.1 if event.delta > 0 else 0.9
        self.tab1_zoom_level *= factor
        self.tab1_zoom_level = max(0.1, min(10.0, self.tab1_zoom_level))
        self.tab1_display_image()

    def tab1_zoom_button(self, factor):
        """Zoom image with +/- keys"""
        if self.tab1_image is None:
            return
        
        self.tab1_zoom_level *= factor
        self.tab1_zoom_level = max(0.1, min(10.0, self.tab1_zoom_level))
        self.tab1_display_image()
    
    def tab1_display_image(self):
        """Display image on canvas"""
        if self.tab1_image is not None:
            display_img = self.tab1_image.copy()

            if len(self.tab1_polygon_points) > 0:
                for i in range(len(self.tab1_polygon_points) - 1):
                    cv2.line(display_img, self.tab1_polygon_points[i],
                            self.tab1_polygon_points[i + 1], (0, 255, 0), 2)

                if self.tab1_polygon_closed:
                    cv2.line(display_img, self.tab1_polygon_points[-1],
                            self.tab1_polygon_points[0], (0, 255, 0), 2)

                for idx, point in enumerate(self.tab1_polygon_points):
                    cv2.circle(display_img, point, 7, (255, 0, 0), -1)
                    cv2.circle(display_img, point, 8, (255, 255, 255), 2)

            pil_img = Image.fromarray(display_img)
            new_width = int(pil_img.width * self.tab1_zoom_level)
            new_height = int(pil_img.height * self.tab1_zoom_level)
            pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            self.tab1_photo = ImageTk.PhotoImage(pil_img)

            self.tab1_canvas.delete("all")
            self.tab1_canvas.create_image(0, 0, anchor='nw', image=self.tab1_photo)
            self.tab1_canvas.configure(scrollregion=self.tab1_canvas.bbox("all"))
            
            
    def tab1_canvas_click(self, event):
        """Handle canvas click to add polygon points or close polygon"""
        if self.tab1_image is not None and not self.tab1_polygon_closed:
            x = int(self.tab1_canvas.canvasx(event.x))
            y = int(self.tab1_canvas.canvasy(event.y))

            # Check if clicking near first point to close polygon
            if len(self.tab1_polygon_points) >= 3:
                first_point = self.tab1_polygon_points[0]
                distance = math.sqrt((x - first_point[0])**2 + (y - first_point[1])**2)

                if distance <= 15:
                    self.tab1_polygon_closed = True
                    self.tab1_display_image()
                    self.tab1_info_label.config(
                        text=f"Polygon closed with {len(self.tab1_polygon_points)} points. Ready to save."
                    )
                    return

            # Add new point
            self.tab1_polygon_points.append((int(x), int(y)))
            self.tab1_display_image()

            if len(self.tab1_polygon_points) < 3:
                self.tab1_info_label.config(
                    text=f"Points selected: {len(self.tab1_polygon_points)} (need at least 3)"
                )
            else:
                self.tab1_info_label.config(
                    text=f"Points selected: {len(self.tab1_polygon_points)} (click on point 1 to close)"
                )

    def tab1_clear_points(self):
        """Clear all polygon points"""
        self.tab1_polygon_points = []
        self.tab1_polygon_closed = False
        self.tab1_display_image()
        self.tab1_info_label.config(
            text="Click on image to select polygon points (click on first point to close)"
        )

    def tab1_save_coordinates(self):
        """Save polygon coordinates to JSON file"""
        if len(self.tab1_polygon_points) < 3:
            messagebox.showwarning("Warning",
                                   "Please select at least 3 points to form a polygon")
            return

        if self.tab1_image_path is None:
            messagebox.showwarning("Warning", "No image loaded")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Coordinates",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if save_path:
            data = {
                "image_path": self.tab1_image_path,
                "image_shape": self.tab1_image.shape[:2],
                "polygon_points": self.tab1_polygon_points
            }

            with open(save_path, 'w') as f:
                json.dump(data, f, indent=4)

            messagebox.showinfo("Success", f"Coordinates saved to {save_path}")

    # ==================== TAB 2: WATERSHED ANALYSIS (Single Image) ====================

    def init_tab2(self):
        """Initialize Single Image tab – Watershed Analysis with new GUI layout."""
        # Variables
        self.tab2_image = None
        self.tab2_cv_image = None
        self.tab2_roi_points = None
        self.tab2_image_path = None
        self.current_step = 0
        self.step_images = {}
        self.step_data = {}
        self.tab2_zoom_level = 1.0

        # Step 1 inline ROI drawing state
        self.step1_polygon_points = []
        self.step1_polygon_closed = False
        self._roi_save_dir = None

        # Default parameters for each step
        self.params = {
            'step_2_rotate_flip': {'enabled': True, 'rotate': 0, 'flip_h': False, 'flip_v': False},
            'step_3_grayscale': {'enabled': True},
            'step_4_sobel': {'enabled': False},
            'step_5_markers': {'enabled': True, 'low_threshold': 90, 'high_threshold': 220},
            'step_6_watershed': {'enabled': False},
            'step_7_morph_close': {'enabled': False, 'iterations': 1},
            'step_8_morph_open': {'enabled': False, 'iterations': 1},
            'step_9_fillholes': {'enabled': False},
            'step_10_labeling': {'enabled': True, 'min_size': 50},
            'step_11_final': {'min_area': 50, 'max_area': 10000},
            'step_12_perspective': {'enabled': True}
        }

        frame = self.tab2  # == self.tab_frames['single']
        frame.configure(bg=self.BG)

        # ── Top toolbar ──────────────────────────────────────────────────
        toolbar = tk.Frame(frame, bg=self.PANEL, height=46)
        toolbar.pack(side='top', fill='x')
        toolbar.pack_propagate(False)

        def _mk_btn(parent, text, cmd):
            return tk.Button(parent, text=text, command=cmd,
                             bg=self.BTN_BG, fg=self.FG, relief='flat',
                             font=('Segoe UI', 9, 'bold'), padx=10, pady=4,
                             activebackground=self.HIGHLIGHT, cursor='hand2')

        _mk_btn(toolbar, 'Load Image',       self.tab2_upload_image    ).pack(side='left', padx=6, pady=7)
        _mk_btn(toolbar, 'Load Config',      self.tab2_load_config     ).pack(side='left', padx=2, pady=7)
        _mk_btn(toolbar, 'Save Config',      self.tab2_save_config     ).pack(side='left', padx=2, pady=7)
        _mk_btn(toolbar, 'Start Processing', self.tab2_start_processing).pack(side='left', padx=2, pady=7)

        self.tab2_info_label = tk.Label(toolbar,
            text="ORIGINAL IMAGE: Raw loaded greyscale image – no processing applied yet",
            bg=self.PANEL, fg=self.FG2, font=('Segoe UI', 9))
        self.tab2_info_label.pack(side='left', padx=14)

        # ── Step number pills [1][2]…[10] ───────────────────────────────
        step_bar = tk.Frame(frame, bg=self.BG)
        step_bar.pack(side='top', fill='x', padx=8, pady=(6, 2))

        self._step_pill_btns = {}
        for i in range(13):
            label = str(i) if i > 0 else '0'
            btn = tk.Button(step_bar, text=label, width=3,
                            bg=self.ACCENT, fg=self.FG2, relief='flat',
                            font=('Segoe UI', 9, 'bold'),
                            activebackground=self.HIGHLIGHT,
                            cursor='hand2',
                            command=lambda s=i: self._jump_to_step(s))
            btn.pack(side='left', padx=2)
            self._step_pill_btns[i] = btn

        # ── Main content (image left, params right) ──────────────────────
        content = tk.Frame(frame, bg=self.BG)
        content.pack(fill='both', expand=True, padx=8, pady=(2, 4))

        # Left: image canvas
        img_frame = tk.Frame(content, bg=self.BG)
        img_frame.pack(side='left', fill='both', expand=True, padx=(0, 6))

        self.tab2_canvas = tk.Canvas(img_frame, bg='#181d22', highlightthickness=0)
        v_sb = ttk.Scrollbar(img_frame, orient='vertical', command=self.tab2_canvas.yview)
        h_sb = ttk.Scrollbar(img_frame, orient='horizontal', command=self.tab2_canvas.xview)
        self.tab2_canvas.configure(yscrollcommand=v_sb.set, xscrollcommand=h_sb.set)
        v_sb.pack(side='right', fill='y')
        h_sb.pack(side='bottom', fill='x')
        self.tab2_canvas.pack(side='left', fill='both', expand=True)

        self.tab2_canvas.bind('<MouseWheel>',
                              lambda e: self.tab2_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self.tab2_canvas.bind('<Shift-MouseWheel>',
                              lambda e: self.tab2_canvas.xview_scroll(int(-1*(e.delta/120)), "units"))
        self.tab2_canvas.bind('<Control-MouseWheel>', self.tab2_zoom)
        # Linux uses Button-4 (scroll up) and Button-5 (scroll down)
        self.tab2_canvas.bind('<Control-Button-4>', self.tab2_zoom)
        self.tab2_canvas.bind('<Control-Button-5>', self.tab2_zoom)
        self.tab2_canvas.bind('<plus>',  lambda e: self.tab2_zoom_button(1.1))
        self.tab2_canvas.bind('<minus>', lambda e: self.tab2_zoom_button(0.9))
        self.tab2_canvas.bind('<equal>', lambda e: self.tab2_zoom_button(1.1))
        self.tab2_canvas.bind('<Button-1>', self.tab2_step1_canvas_click)
        # Ctrl+drag to zoom
        self.tab2_canvas.bind('<Control-ButtonPress-1>',  self._tab2_drag_zoom_start)
        self.tab2_canvas.bind('<Control-B1-Motion>',      self._tab2_drag_zoom_motion)

        # Placeholder label when no image loaded
        self._img_placeholder = tk.Label(self.tab2_canvas,
            text='▲\nLoad Image to begin',
            bg='#181d22', fg='#3e4e5e',
            font=('Segoe UI', 14), justify='center')
        self._img_placeholder.place(relx=0.5, rely=0.5, anchor='center')

        # Right: parameters panel
        right_panel = tk.Frame(content, bg=self.PANEL, width=320)
        right_panel.pack(side='right', fill='y')
        right_panel.pack_propagate(False)

        # Step info inside right panel
        step_info_frame = tk.Frame(right_panel, bg=self.PANEL)
        step_info_frame.pack(fill='x', padx=10, pady=(12, 4))

        self.step_name_label = tk.Label(step_info_frame, text="Step 0: Original Image",
                                        bg=self.PANEL, fg=self.HIGHLIGHT,
                                        font=('Segoe UI', 11, 'bold'), wraplength=280, justify='left')
        self.step_name_label.pack(anchor='w')

        self.step_desc_label = tk.Label(step_info_frame, text="Load an image to begin",
                                        bg=self.PANEL, fg=self.FG2,
                                        font=('Segoe UI', 9), wraplength=280, justify='left')
        self.step_desc_label.pack(anchor='w', pady=(4, 0))

        sep = tk.Frame(right_panel, bg=self.ACCENT, height=1)
        sep.pack(fill='x', padx=10, pady=8)

        # Parameters header
        tk.Label(right_panel, text='Parameters',
                 bg=self.PANEL, fg=self.FG,
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=12)

        # Scrollable params canvas
        params_outer = tk.Frame(right_panel, bg=self.PANEL)
        params_outer.pack(fill='both', expand=True, padx=6, pady=6)

        self.params_canvas = tk.Canvas(params_outer, bg=self.PANEL, highlightthickness=0)
        params_scrollbar = ttk.Scrollbar(params_outer, orient='vertical',
                                         command=self.params_canvas.yview)
        self.params_frame = tk.Frame(self.params_canvas, bg=self.PANEL)

        self.params_canvas.configure(yscrollcommand=params_scrollbar.set)
        params_scrollbar.pack(side='right', fill='y')
        self.params_canvas.pack(side='left', fill='both', expand=True)

        self.params_canvas_window = self.params_canvas.create_window(
            (0, 0), window=self.params_frame, anchor='nw')

        self.params_frame.bind('<Configure>',
                               lambda e: self.params_canvas.configure(
                                   scrollregion=self.params_canvas.bbox('all')))

        # ── Bottom navigation (Prev / Next) ─────────────────────────────
        nav_bar = tk.Frame(frame, bg=self.PANEL, height=44)
        nav_bar.pack(side='bottom', fill='x')
        nav_bar.pack_propagate(False)

        self.progress_label = tk.Label(nav_bar, text="Step 0 / 12",
                                       bg=self.PANEL, fg=self.FG2,
                                       font=('Segoe UI', 10))
        self.progress_label.pack(side='left', padx=16)

        # ── Zoom controls in nav bar ─────────────────────────────────────
        zoom_frame = tk.Frame(nav_bar, bg=self.PANEL)
        zoom_frame.pack(side='left', padx=20)

        tk.Label(zoom_frame, text='Zoom:', bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9)).pack(side='left', padx=(0, 4))

        tk.Button(zoom_frame, text='−', width=3,
                  command=lambda: self.tab2_zoom_button(0.8),
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 11, 'bold'), pady=2,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='left', padx=2)

        tk.Button(zoom_frame, text='+', width=3,
                  command=lambda: self.tab2_zoom_button(1.25),
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 11, 'bold'), pady=2,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='left', padx=2)

        tk.Button(zoom_frame, text='Reset', width=5,
                  command=self.tab2_zoom_reset,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), pady=2,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='left', padx=2)

        self.zoom_label = tk.Label(zoom_frame, text='100%', bg=self.PANEL, fg=self.FG2,
                                   font=('Segoe UI', 9), width=5)
        self.zoom_label.pack(side='left', padx=(4, 0))

        self.next_button = tk.Button(nav_bar, text='Next  ›',
                                     command=self.tab2_next_step,
                                     bg=self.HIGHLIGHT, fg='#ffffff', relief='flat',
                                     font=('Segoe UI', 10, 'bold'), padx=14, pady=6,
                                     activebackground=self.BTN_ACT, cursor='hand2',
                                     state='disabled')
        self.next_button.pack(side='right', padx=10, pady=6)

        self.prev_button = tk.Button(nav_bar, text='‹  Previous',
                                     command=self.tab2_previous_step,
                                     bg=self.BTN_BG, fg=self.FG, relief='flat',
                                     font=('Segoe UI', 10, 'bold'), padx=14, pady=6,
                                     activebackground=self.HIGHLIGHT, cursor='hand2',
                                     state='disabled')
        self.prev_button.pack(side='right', padx=4, pady=6)

        self.tab2_photo = None

    def _jump_to_step(self, step):
        """Jump to a specific step when a pill button is clicked."""
        if step in self.step_images:
            self.current_step = step
            self.display_current_step()

    def _update_step_pills(self):
        """Highlight the currently active step pill."""
        for i, btn in self._step_pill_btns.items():
            if i == self.current_step:
                btn.configure(bg=self.HIGHLIGHT, fg='#ffffff')
            elif i in self.step_images or (i == 1 and len(self.step1_polygon_points) > 0):
                btn.configure(bg=self.ACCENT, fg=self.FG)
            else:
                btn.configure(bg=self.ACCENT, fg=self.FG2)

    def tab2_upload_image(self):
        """Upload image for watershed analysis"""
        file_path = filedialog.askopenfilename(
            title="Select Image",
            filetypes=[("Image files", "*.jpg *.jpeg *.png"),
                       ("All files", "*.*")]
        )

        if file_path:
            self.tab2_image_path = file_path
            self.tab2_cv_image = cv2.imread(file_path)
            self.tab2_image = cv2.cvtColor(self.tab2_cv_image, cv2.COLOR_BGR2RGB)

            # Display the image immediately
            self.step_images = {}
            self.step_images[0] = self.tab2_image.copy()
            self.current_step = 0
            self.display_current_step()

            self.tab2_info_label.config(text="Image loaded. Load ROI coordinates and click Start Processing.")

    def tab2_load_roi(self):
        """Load ROI coordinates from JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select ROI Coordinates File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            with open(file_path, 'r') as f:
                data = json.load(f)

            self.tab2_roi_points = np.array(data['polygon_points'], dtype=np.int32)

            # Load parameters if they exist in the file
            if 'pipeline_params' in data:
                self.params = data['pipeline_params']

            self.tab2_info_label.config(
                text=f"ROI loaded with {len(self.tab2_roi_points)} points. Click Start Processing."
            )

    def tab2_zoom(self, event):
        """Zoom image with Ctrl+MouseWheel (Windows/macOS: delta; Linux: num)"""
        if self.tab2_image is None:
            return
        # event.delta works on Windows/macOS; on Linux Button-4/5 give delta=0
        if event.delta != 0:
            factor = 1.1 if event.delta > 0 else 0.9
        else:
            # Linux scroll up = Button-4, scroll down = Button-5
            factor = 1.1 if getattr(event, 'num', 0) == 4 else 0.9
        self.tab2_zoom_level *= factor
        self.tab2_zoom_level = max(0.1, min(10.0, self.tab2_zoom_level))
        self._tab2_update_zoom_label()
        if self.current_step == 1:
            self.display_step1_drawing()
        elif self.current_step in self.step_images:
            self.display_current_step()

    def tab2_zoom_button(self, factor):
        """Zoom image with +/- buttons or keys"""
        if self.current_step != 1 and self.current_step not in self.step_images:
            return
        self.tab2_zoom_level *= factor
        self.tab2_zoom_level = max(0.1, min(10.0, self.tab2_zoom_level))
        self._tab2_update_zoom_label()
        self.display_current_step()
        
    def tab2_zoom_reset(self):
        """Reset zoom to 100%."""
        self.tab2_zoom_level = 1.0
        self._tab2_update_zoom_label()
        self.display_current_step()

    def _tab2_update_zoom_label(self):
        """Refresh the zoom % label in the nav bar."""
        if hasattr(self, 'zoom_label'):
            self.zoom_label.config(text=f"{int(self.tab2_zoom_level * 100)}%")

    def _tab2_drag_zoom_start(self, event):
        """Record start Y position for Ctrl+drag zoom."""
        self._drag_zoom_start_y = event.y
        self._drag_zoom_start_level = self.tab2_zoom_level

    def _tab2_drag_zoom_motion(self, event):
        """Ctrl+drag up = zoom in, drag down = zoom out."""
        if not hasattr(self, '_drag_zoom_start_y'):
            return
        delta = self._drag_zoom_start_y - event.y   # positive = dragged up
        factor = 1.0 + delta * 0.005                # 0.5% per pixel
        new_level = self._drag_zoom_start_level * factor
        self.tab2_zoom_level = max(0.1, min(10.0, new_level))
        self._tab2_update_zoom_label()
        if self.current_step != 1 and self.current_step not in self.step_images:
            return
        self.display_current_step()

    def tab2_step1_canvas_click(self, event):
        """Handle canvas click: only active during Step 1 ROI drawing."""
        if self.current_step != 1:
            return
        if self.tab2_image is None:
            return
        if self.step1_polygon_closed:
            return

        # Convert canvas coords → image coords (account for zoom)
        cx = int(self.tab2_canvas.canvasx(event.x))
        cy = int(self.tab2_canvas.canvasy(event.y))
        x = int(cx / self.tab2_zoom_level)
        y = int(cy / self.tab2_zoom_level)

        # Clamp to image bounds
        h, w = self.tab2_image.shape[:2]
        x = max(0, min(w - 1, x))
        y = max(0, min(h - 1, y))

        # Check if clicking near first point to close polygon
        if len(self.step1_polygon_points) >= 3:
            fp = self.step1_polygon_points[0]
            dist = math.sqrt((x - fp[0])**2 + (y - fp[1])**2)
            if dist <= int(15 / self.tab2_zoom_level):
                self.step1_polygon_closed = True
                self.tab2_roi_points = np.array(self.step1_polygon_points, dtype=np.int32)
                self._step1_auto_save()
                self._display_step1_drawing()
                return

        self.step1_polygon_points.append((x, y))
        self._display_step1_drawing()

    def _step1_auto_save(self):
        """Auto-save ROI coordinates JSON next to the loaded image.
           The config JSON (Save Config) will use the same folder."""
        if self.tab2_image_path is None:
            return
        img_dir  = os.path.dirname(os.path.abspath(self.tab2_image_path))
        img_stem = Path(self.tab2_image_path).stem
        roi_path = os.path.join(img_dir, f"{img_stem}_roi.json")

        data = {
            "image_path":  self.tab2_image_path,
            "image_shape": list(self.tab2_image.shape[:2]),
            "polygon_points": self.step1_polygon_points,
            "pipeline_params": self.params
        }
        with open(roi_path, 'w') as f:
            json.dump(data, f, indent=4)

        # ── Also write a copy to configurations/conf.json so Batch tab
        #    can auto-load it on next startup ──────────────────────────
        try:
            conf_dir = Path(__file__).resolve().parent / 'configurations'
            conf_dir.mkdir(exist_ok=True)
            with open(conf_dir / 'conf.json', 'w') as _cf:
                json.dump(data, _cf, indent=4)
        except Exception:
            pass  # never block the main save if this fails

        # Remember save folder so Save Config lands in same place
        self._roi_save_dir = img_dir
        self.tab2_info_label.config(
            text=f"ROI auto-saved → {roi_path}  |  Click Next to run pipeline.")
        messagebox.showinfo("ROI Saved",
            f"Polygon closed!\n\nCoordinates auto-saved to:\n{roi_path}\n\nClick Next to run the pipeline.")

    def tab2_step1_clear_points(self):
        """Clear Step 1 polygon drawing."""
        self.step1_polygon_points = []
        self.step1_polygon_closed = False
        self.tab2_roi_points = None
        self._display_step1_drawing()

    def tab2_start_processing(self):
        """Start the step-by-step watershed processing.

        Behaviour:
        - If the user has previously clicked 'Load Config' and the config
          contained polygon_points, those points are already stored in
          self.step1_polygon_points / self.step1_polygon_closed.
          → Step 1 will open with the saved polygon pre-drawn so the user
            can verify or modify it before proceeding.
        - If no config was loaded (or the config had no polygon), the Step 1
          polygon state is empty and the user draws manually as before.
        """
        if self.tab2_image is None:
            messagebox.showwarning("Warning", "Please upload an image first")
            return

        # ── Determine whether a pre-loaded ROI is available ──────────────
        # (populated by tab2_load_config when JSON contains polygon_points)
        roi_pre_loaded = (
            self.step1_polygon_closed and
            len(self.step1_polygon_points) >= 3
        )

        # If no config was loaded at all, make sure the drawing state is clean
        if not roi_pre_loaded:
            self.tab2_roi_points      = None
            self.step1_polygon_points = []
            self.step1_polygon_closed = False

        # Show Step 0: original image only
        self.step_images = {}
        self.step_images[0] = self.tab2_image.copy()
        self.current_step = 0
        self.display_current_step()

        # Allow moving to Step 1 (ROI draw/review)
        self.next_button.config(state='normal')
        self.prev_button.config(state='disabled')

        if roi_pre_loaded:
            self.tab2_info_label.config(
                text=f"✔ ROI from loaded config pre-filled "
                     f"({len(self.step1_polygon_points)} points). "
                     f"Click Next to review / edit in Step 1, then proceed.")
        else:
            self.tab2_info_label.config(
                text="Step 0 loaded. Click Next to draw ROI polygon in Step 1.")

    def process_all_steps(self):
        """Process all watershed steps using Scikit-Image approach (Sobel + Histogram Markers)"""
        self.step_images = {}
        self.step_data = {}

        # Step 0: Original Image
        self.step_images[0] = self.tab2_image.copy()

        # Step 1: RoI Masked Image (4-point polygon)
        mask = np.zeros(self.tab2_image.shape[:2], dtype=np.uint8)
        cv2.fillPoly(mask, [self.tab2_roi_points], 255)
        masked_image = self.tab2_cv_image.copy()
        masked_image[mask == 0] = [0, 0, 0]
        self.step_images[1] = cv2.cvtColor(masked_image, cv2.COLOR_BGR2RGB)
        self.step_data['mask'] = mask
        self.step_data['masked_bgr'] = masked_image

        # Step 2: Rotate / Flip  (NEW)
        rotated_bgr = masked_image.copy()
        if self.params['step_2_rotate_flip']['enabled']:
            angle = self.params['step_2_rotate_flip']['rotate']
            if angle != 0:
                h, w = rotated_bgr.shape[:2]
                cx, cy = w // 2, h // 2
                M_rot = cv2.getRotationMatrix2D((cx, cy), -angle, 1.0)
                rotated_bgr = cv2.warpAffine(rotated_bgr, M_rot, (w, h),
                                             borderMode=cv2.BORDER_CONSTANT,
                                             borderValue=(0, 0, 0))
            if self.params['step_2_rotate_flip']['flip_h']:
                rotated_bgr = cv2.flip(rotated_bgr, 1)
            if self.params['step_2_rotate_flip']['flip_v']:
                rotated_bgr = cv2.flip(rotated_bgr, 0)
        self.step_images[2] = cv2.cvtColor(rotated_bgr, cv2.COLOR_BGR2RGB)
        self.step_data['rotated_bgr'] = rotated_bgr

        # Step 3: Grayscale Conversion  (was Step 2)
        if self.params['step_3_grayscale']['enabled']:
            gray = cv2.cvtColor(rotated_bgr, cv2.COLOR_BGR2GRAY)
            self.step_images[3] = cv2.cvtColor(gray, cv2.COLOR_GRAY2RGB)
        else:
            gray = cv2.cvtColor(rotated_bgr, cv2.COLOR_BGR2GRAY)
            self.step_images[3] = self.step_images[2].copy()
        self.step_data['gray'] = gray

        # Step 4: Sobel Gradient (Elevation Map)  (was Step 3)
        if self.params['step_4_sobel']['enabled']:
            sobel_x = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
            sobel_y = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
            elevation_map = np.hypot(sobel_x, sobel_y)
            elevation_display = cv2.normalize(elevation_map, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
        else:
            elevation_map = gray.astype(np.float64)
            elevation_display = gray
        self.step_images[4] = cv2.cvtColor(elevation_display, cv2.COLOR_GRAY2RGB)
        self.step_data['elevation_map'] = elevation_map

        # Step 5: Histogram-based Markers  (was Step 4)
        if self.params['step_5_markers']['enabled']:
            low_thresh  = self.params['step_5_markers']['low_threshold']
            high_thresh = self.params['step_5_markers']['high_threshold']

            inv = 255 - gray
            background = cv2.GaussianBlur(inv, (0, 0), sigmaX=50, sigmaY=50)
            gbImg = cv2.subtract(inv, background)
            _, gb = cv2.threshold(gbImg, 10, 255, cv2.THRESH_BINARY)

            markers = np.zeros_like(gray, dtype=np.int32)
            background_mask = gray < low_thresh
            markers[background_mask] = gb[background_mask]

            foreground_mask = gray > high_thresh
            markers[foreground_mask] = gray[foreground_mask].astype(np.int32) + 1

            markers_display = np.zeros_like(gray)
            markers_display[background_mask] = gb[background_mask]
            markers_display[foreground_mask] = gray[foreground_mask]
        else:
            markers = np.zeros_like(gray, dtype=np.int32)
            markers[gray > 0] = gray[gray > 0].astype(np.int32) + 1
            markers_display = gray.copy()

        self.step_images[5] = cv2.cvtColor(markers_display, cv2.COLOR_GRAY2RGB)
        self.step_data['markers'] = markers

        # Step 6: Watershed Application  (was Step 5)
        if self.params['step_6_watershed']['enabled']:
            elevation_map_uint8 = elevation_display
            elevation_map_color = cv2.cvtColor(elevation_map_uint8, cv2.COLOR_GRAY2BGR)
            markers_watershed = markers.copy()
            segmentation = cv2.watershed(elevation_map_color, markers_watershed)

            watershed_display = np.zeros_like(self.step_data['rotated_bgr'])
            watershed_display[segmentation == 1]  = [50, 50, 50]
            watershed_display[segmentation == 2]  = [0, 200, 0]
            watershed_display[segmentation == -1] = [255, 0, 0]
            watershed_display[segmentation == 0]  = [0, 0, 0]
        else:
            segmentation = markers.copy()
            watershed_display = self.step_data['rotated_bgr'].copy()

        self.step_images[6] = cv2.cvtColor(watershed_display, cv2.COLOR_BGR2RGB)
        self.step_data['segmentation'] = segmentation

        # Convert segmentation to binary
        if self.params['step_5_markers']['enabled']:
            high_thresh = self.params['step_5_markers']['high_threshold']
            binary_seg = (segmentation > high_thresh).astype(np.uint8)
        else:
            binary_seg = (segmentation > 0).astype(np.uint8)

        # Step 7: Morphological Close  (was Step 6)
        if self.params['step_7_morph_close']['enabled']:
            iterations = self.params['step_7_morph_close']['iterations']
            kernel = np.ones((3, 3), np.uint8)
            morph_close = cv2.morphologyEx(binary_seg, cv2.MORPH_CLOSE, kernel, iterations=iterations)
        else:
            morph_close = binary_seg.copy()

        self.step_images[7] = cv2.cvtColor(morph_close * 255, cv2.COLOR_GRAY2RGB)
        self.step_data['morph_close'] = morph_close

        # Step 8: Morphological Open  (was Step 7)
        if self.params['step_8_morph_open']['enabled']:
            iterations = self.params['step_8_morph_open']['iterations']
            kernel = np.ones((3, 3), np.uint8)
            morph_open = cv2.morphologyEx(morph_close, cv2.MORPH_OPEN, kernel, iterations=iterations)
        else:
            morph_open = morph_close.copy()

        self.step_images[8] = cv2.cvtColor(morph_open * 255, cv2.COLOR_GRAY2RGB)
        self.step_data['morph_open'] = morph_open

        # Step 9: Fill Holes  (was Step 8)
        if self.params['step_9_fillholes']['enabled']:
            filled = ndi.binary_fill_holes(morph_open).astype(np.uint8) * 255
        else:
            filled = morph_open * 255

        self.step_images[9] = cv2.cvtColor(filled, cv2.COLOR_GRAY2RGB)
        self.step_data['filled'] = filled

        # Step 10: Label Connected Components  (was Step 9)
        if self.params['step_10_labeling']['enabled']:
            labeled_array, num_features = ndi.label(filled)
            min_size = self.params['step_10_labeling']['min_size']
            sizes = np.bincount(labeled_array.ravel())
            mask_sizes = sizes > min_size
            mask_sizes[0] = 0
            cleaned_labels = mask_sizes[labeled_array]
            labeled_cleaned, num_cleaned = ndi.label(cleaned_labels)
            labeled_display = cv2.normalize(labeled_cleaned, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
            labeled_colored = cv2.applyColorMap(labeled_display, cv2.COLORMAP_JET)
        else:
            labeled_cleaned, num_cleaned = ndi.label(filled)
            labeled_display = cv2.normalize(labeled_cleaned, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
            labeled_colored = cv2.applyColorMap(labeled_display, cv2.COLORMAP_JET)

        self.step_images[10] = cv2.cvtColor(labeled_colored, cv2.COLOR_BGR2RGB)
        self.step_data['labeled'] = labeled_cleaned
        self.step_data['num_labels'] = num_cleaned

        # Step 11: Final Result with Colored Segments  (was Step 10)
        result = self.tab2_cv_image.copy()
        min_area = self.params['step_11_final']['min_area']
        max_area = self.params['step_11_final']['max_area']
        segment_color = (0, 0, 255)

        for region_label in range(1, self.step_data['num_labels'] + 1):
            region_mask = (labeled_cleaned == region_label).astype(np.uint8) * 255
            contours, _ = cv2.findContours(region_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if contours:
                for contour in contours:
                    area = cv2.contourArea(contour)
                    if min_area <= area <= max_area:
                        cv2.drawContours(result, [contour], -1, segment_color, -1)

        self.step_images[11] = cv2.cvtColor(result, cv2.COLOR_BGR2RGB)

        # Step 12: Perspective Warp  (was Step 11)
        if self.params['step_12_perspective']['enabled'] and self.tab2_roi_points is not None:
            roi_pts = self.tab2_roi_points.reshape(-1, 2).astype(np.float32)
            s = roi_pts.sum(axis=1)
            d = np.diff(roi_pts, axis=1).ravel()
            tl = roi_pts[np.argmin(s)]
            br = roi_pts[np.argmax(s)]
            tr = roi_pts[np.argmin(d)]
            bl = roi_pts[np.argmax(d)]
            src_quad = np.array([tl, tr, br, bl], dtype=np.float32)
            out_w = 272*8
            out_h = 272*8
            dst_quad = np.array([[0, 0], [out_w - 1, 0],
                                 [out_w - 1, out_h - 1], [0, out_h - 1]], dtype=np.float32)
            M = cv2.getPerspectiveTransform(src_quad, dst_quad)
            warped = cv2.warpPerspective(result, M, (out_w, out_h),
                                         flags=cv2.INTER_LINEAR,
                                         borderMode=cv2.BORDER_CONSTANT,
                                         borderValue=(0, 0, 0))
        else:
            warped = result.copy()

        self.step_images[12] = cv2.cvtColor(warped, cv2.COLOR_BGR2RGB)
        self.step_data['warped'] = warped

    def display_current_step(self):
        """Display the current step's image and parameters"""
        # Step 1 is interactive: display it even before step_images[1] exists
        if self.current_step == 1:
            self._display_step1_drawing()
            return

        if self.current_step not in self.step_images:
            return

        # Update step info
        step_names = [
            "Step 0: Original Image",
            "Step 1: RoI Selection (4-Point Polygon)",
            "Step 2: Rotate / Flip",
            "Step 3: Grayscale Conversion",
            "Step 4: Sobel Gradient (Elevation Map)",
            "Step 5: Histogram Markers",
            "Step 6: Watershed Application",
            "Step 7: Morphological Close",
            "Step 8: Morphological Open",
            "Step 9: Fill Holes",
            "Step 10: Connected Component Labeling",
            "Step 11: Final Result",
            "Step 12: Perspective Warp"
        ]

        step_descriptions = [
            "Original loaded image without any processing.",
            "Image with Area of Interest applied. Regions outside the 4-point polygon are blacked out.",
            "Rotate and/or flip the masked image before further processing.",
            "Image converted to grayscale for processing.",
            "Sobel gradient (elevation map) computed. High gradient values form barriers between regions.",
            "Markers with varying values: background uses low intensities, foreground uses high intensities.",
            "Watershed algorithm floods elevation map from markers. Boundaries marked in red.",
            "Morphological closing to fill small holes and connect nearby objects.",
            "Morphological opening to remove small noise and separate touching objects.",
            "Binary fill holes removes small holes in detected foreground regions.",
            "Connected component labeling identifies individual objects. Small objects filtered out.",
            "Final result with detected segments colored in red.",
            "Perspective warp: ROI polygon corners mapped to a flat rectangle for top-down indicator view."
        ]

        self.step_name_label.config(text=step_names[self.current_step])
        self.step_desc_label.config(text=step_descriptions[self.current_step])
        self.progress_label.config(text=f"Step {self.current_step} / 12")

        # Hide placeholder
        self._img_placeholder.place_forget()

        # Display image
        image = self.step_images[self.current_step]
        pil_img = Image.fromarray(image)
        new_width = int(pil_img.width * self.tab2_zoom_level)
        new_height = int(pil_img.height * self.tab2_zoom_level)
        pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        self.tab2_photo = ImageTk.PhotoImage(pil_img)

        self.tab2_canvas.delete("all")
        self.tab2_canvas.create_image(0, 0, anchor='nw', image=self.tab2_photo)
        self.tab2_canvas.configure(scrollregion=self.tab2_canvas.bbox("all"))

        # Update parameter controls
        self.update_parameter_controls()

        # Update navigation buttons
        self.prev_button.config(state='normal' if self.current_step > 0 else 'disabled')
        self.next_button.config(state='normal' if self.current_step < 12 else 'disabled')

        # Update step pills
        self._update_step_pills()

    def _display_step1_drawing(self):
        """Render Step 1 interactive ROI drawing on the main tab2 canvas."""
        step_names = ["Step 0: Original Image",
                      "Step 1: RoI Selection (Draw Polygon)",]
        self.step_name_label.config(text="Step 1: RoI Selection (Draw Polygon)")
        self.step_desc_label.config(
            text="Click on the image to place polygon points. "
                 "Click near the first point (green circle) to close the polygon.")
        self.progress_label.config(text="Step 1 / 12")
        self._img_placeholder.place_forget()

        # Draw original image with polygon overlay
        display_img = self.tab2_image.copy()
        pts = self.step1_polygon_points
        zoom = self.tab2_zoom_level

        if len(pts) > 0:
            # Draw lines between consecutive points
            for i in range(len(pts) - 1):
                p1 = (int(pts[i][0] * zoom), int(pts[i][1] * zoom))
                p2 = (int(pts[i+1][0] * zoom), int(pts[i+1][1] * zoom))
                cv2.line(display_img, pts[i], pts[i+1], (0, 255, 0), 2)
            if self.step1_polygon_closed:
                cv2.line(display_img, pts[-1], pts[0], (0, 255, 0), 2)
                # Fill polygon overlay semi-transparent style (solid tint)
                overlay = display_img.copy()
                pts_arr = np.array(pts, dtype=np.int32)
                cv2.fillPoly(overlay, [pts_arr], (0, 255, 0))
                cv2.addWeighted(overlay, 0.15, display_img, 0.85, 0, display_img)
                cv2.polylines(display_img, [pts_arr], True, (0, 255, 0), 2)
            # Draw point circles
            for idx, pt in enumerate(pts):
                color = (255, 215, 0) if idx == 0 else (255, 0, 0)
                cv2.circle(display_img, pt, 7, color, -1)
                cv2.circle(display_img, pt, 8, (255, 255, 255), 2)
                cv2.putText(display_img, str(idx + 1), (pt[0]+10, pt[1]-6),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)

        pil_img = Image.fromarray(display_img)
        new_width  = int(pil_img.width  * self.tab2_zoom_level)
        new_height = int(pil_img.height * self.tab2_zoom_level)
        pil_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
        self.tab2_photo = ImageTk.PhotoImage(pil_img)

        self.tab2_canvas.delete("all")
        self.tab2_canvas.create_image(0, 0, anchor='nw', image=self.tab2_photo)
        self.tab2_canvas.configure(scrollregion=self.tab2_canvas.bbox("all"))

        # Update right panel controls for step 1
        self.update_parameter_controls()

        # Nav buttons
        self.prev_button.config(state='normal')
        self.next_button.config(state='normal')
        self._update_step_pills()

        # Update info label
        if self.step1_polygon_closed:
            self.tab2_info_label.config(
                text=f"Polygon closed with {len(pts)} points. "
                     f"Coordinates auto-saved. Click Next to run pipeline.")
        elif len(pts) == 0:
            self.tab2_info_label.config(
                text="Step 1: Click on the image to place ROI polygon points.")
        elif len(pts) < 3:
            self.tab2_info_label.config(
                text=f"Points placed: {len(pts)}. Need at least 3 points.")
        else:
            self.tab2_info_label.config(
                text=f"Points placed: {len(pts)}. Click near point 1 (gold circle) to close polygon.")

    def update_parameter_controls(self):
        """Update parameter controls based on current step"""
        # Clear existing controls
        for widget in self.params_frame.winfo_children():
            widget.destroy()

        step = self.current_step

        if step == 1:   # ROI Drawing
            self.create_step1_controls()
        elif step == 2:   # Rotate / Flip (NEW)
            self.create_step2_controls()
        elif step == 3:  # Grayscale
            self.create_step3_controls()
        elif step == 4:  # Sobel
            self.create_step4_controls()
        elif step == 5:  # Markers
            self.create_step5_controls()
        elif step == 6:  # Watershed
            self.create_step6_controls()
        elif step == 7:  # Morphological Close
            self.create_step7_controls()
        elif step == 8:  # Morphological Open
            self.create_step8_controls()
        elif step == 9:  # Fill Holes
            self.create_step9_controls()
        elif step == 10:  # Labeling
            self.create_step10_controls()
        elif step == 11:  # Final
            self.create_step11_controls()
        elif step == 12:  # Perspective Warp
            self.create_step12_controls()
        else:
            ttk.Label(self.params_frame, text="No adjustable parameters for this step.",
                     wraplength=280).pack(pady=10)

    def create_step1_controls(self):
        """Right-panel controls for Step 1: ROI Drawing."""
        tk.Label(self.params_frame,
                 text="Draw ROI Polygon",
                 bg=self.PANEL, fg=self.HIGHLIGHT,
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(8, 4))

        tk.Label(self.params_frame,
                 text="1. Click on the image to place polygon points.\n"
                      "2. Point 1 is shown as a gold circle.\n"
                      "3. Click near point 1 to close the polygon.\n"
                      "4. Coordinates are auto-saved next to your image.\n"
                      "5. Click Next to run Steps 2-12.",
                 bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), wraplength=280, justify='left').pack(anchor='w', padx=4, pady=(0, 12))

        tk.Frame(self.params_frame, bg=self.ACCENT, height=1).pack(fill='x', pady=6)

        # Points counter
        n = len(self.step1_polygon_points)
        status = "Closed ✔" if self.step1_polygon_closed else f"{n} point(s) placed"
        self._step1_count_lbl = tk.Label(self.params_frame,
                 text=f"Status: {status}",
                 bg=self.PANEL, fg=self.FG,
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=4, pady=(0, 8))

        tk.Button(self.params_frame,
                  text="Clear Points",
                  command=self.tab2_step1_clear_points,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 9, 'bold'), padx=10, pady=4,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack(anchor='w', padx=4)

    def create_step2_controls(self):
        """Create controls for Step 2: Rotate / Flip"""
        skip_var = tk.BooleanVar(value=not self.params['step_2_rotate_flip']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_2_rotate_flip', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)

        # Rotation angle
        ttk.Label(self.params_frame, text="Rotation Angle (°):").pack(anchor='w', pady=(5, 2))
        rot_frame = ttk.Frame(self.params_frame)
        rot_frame.pack(fill='x', pady=2)

        rot_var = tk.IntVar(value=self.params['step_2_rotate_flip']['rotate'])
        rot_scale = ttk.Scale(rot_frame, from_=-180, to=180, orient='horizontal',
                              variable=rot_var,
                              command=lambda v: self.update_step2_rotate(rot_var))
        rot_scale.pack(side='left', fill='x', expand=True, padx=(0, 5))
        rot_spinbox = ttk.Spinbox(rot_frame, from_=-180, to=180, textvariable=rot_var,
                                  width=6, command=lambda: self.update_step2_rotate(rot_var))
        rot_spinbox.pack(side='right')
        rot_spinbox.bind('<Return>', lambda e: self.update_step2_rotate(rot_var))
        rot_spinbox.bind('<FocusOut>', lambda e: self.update_step2_rotate(rot_var))

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)

        # Flip horizontal
        flip_h_var = tk.BooleanVar(value=self.params['step_2_rotate_flip']['flip_h'])
        ttk.Checkbutton(self.params_frame, text="Flip Horizontal (left ↔ right)",
                        variable=flip_h_var,
                        command=lambda: self.update_step2_flip(flip_h_var, None)).pack(anchor='w', pady=3)

        # Flip vertical
        flip_v_var = tk.BooleanVar(value=self.params['step_2_rotate_flip']['flip_v'])
        ttk.Checkbutton(self.params_frame, text="Flip Vertical (top ↔ bottom)",
                        variable=flip_v_var,
                        command=lambda: self.update_step2_flip(None, flip_v_var)).pack(anchor='w', pady=3)

        self.param_widgets = {'rot_var': rot_var, 'flip_h_var': flip_h_var, 'flip_v_var': flip_v_var}

    def create_step3_controls(self):
        """Create controls for Step 3: Grayscale Conversion"""
        skip_var = tk.BooleanVar(value=not self.params['step_3_grayscale']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_3_grayscale', skip_var)).pack(anchor='w', pady=5)

        ttk.Label(self.params_frame, text="Note: Skipping shows RGB but processes grayscale internally.",
                 wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=10)

    def create_step4_controls(self):
        """Create controls for Step 4: Sobel Gradient"""
        skip_var = tk.BooleanVar(value=not self.params['step_4_sobel']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_4_sobel', skip_var)).pack(anchor='w', pady=5)

        ttk.Label(self.params_frame, text="Note: Sobel computes image gradient as elevation map for watershed. High gradients form barriers between regions.",
                 wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=10)

    def create_step5_controls(self):
        """Create controls for Step 5: Histogram Markers"""
        skip_var = tk.BooleanVar(value=not self.params['step_5_markers']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_5_markers', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)

        # Low threshold
        ttk.Label(self.params_frame, text="Low Threshold (Background):").pack(anchor='w', pady=(5, 2))
        low_frame = ttk.Frame(self.params_frame)
        low_frame.pack(fill='x', pady=2)
        low_var = tk.IntVar(value=self.params['step_5_markers']['low_threshold'])
        ttk.Scale(low_frame, from_=0, to=255, orient='horizontal', variable=low_var,
                  command=lambda v: self.update_step5_params(low_var, None)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        low_spinbox = ttk.Spinbox(low_frame, from_=0, to=255, textvariable=low_var,
                                  width=6, command=lambda: self.update_step5_params(low_var, None))
        low_spinbox.pack(side='right')
        low_spinbox.bind('<Return>', lambda e: self.update_step5_params(low_var, None))
        low_spinbox.bind('<FocusOut>', lambda e: self.update_step5_params(low_var, None))

        # High threshold
        ttk.Label(self.params_frame, text="High Threshold (Foreground):").pack(anchor='w', pady=(10, 2))
        high_frame = ttk.Frame(self.params_frame)
        high_frame.pack(fill='x', pady=2)
        high_var = tk.IntVar(value=self.params['step_5_markers']['high_threshold'])
        ttk.Scale(high_frame, from_=0, to=255, orient='horizontal', variable=high_var,
                  command=lambda v: self.update_step5_params(None, high_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        high_spinbox = ttk.Spinbox(high_frame, from_=0, to=255, textvariable=high_var,
                                   width=6, command=lambda: self.update_step5_params(None, high_var))
        high_spinbox.pack(side='right')
        high_spinbox.bind('<Return>', lambda e: self.update_step5_params(None, high_var))
        high_spinbox.bind('<FocusOut>', lambda e: self.update_step5_params(None, high_var))
        self.param_widgets = {'low_var': low_var, 'high_var': high_var}

    def create_step6_controls(self):
        """Create controls for Step 6: Watershed"""
        skip_var = tk.BooleanVar(value=not self.params['step_6_watershed']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_6_watershed', skip_var)).pack(anchor='w', pady=5)

        ttk.Label(self.params_frame, text="Note: Watershed floods elevation map from markers to segment regions.",
                 wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=10)

    def create_step7_controls(self):
        """Create controls for Step 7: Morphological Close"""
        skip_var = tk.BooleanVar(value=not self.params['step_7_morph_close']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_7_morph_close', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(self.params_frame, text="Number of Iterations:").pack(anchor='w', pady=(5, 2))
        iter_frame = ttk.Frame(self.params_frame)
        iter_frame.pack(fill='x', pady=2)
        iter_var = tk.IntVar(value=self.params['step_7_morph_close']['iterations'])
        ttk.Scale(iter_frame, from_=1, to=10, orient='horizontal', variable=iter_var,
                  command=lambda v: self.update_step7_iterations(iter_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        iter_spinbox = ttk.Spinbox(iter_frame, from_=1, to=10, textvariable=iter_var,
                                   width=6, command=lambda: self.update_step7_iterations(iter_var))
        iter_spinbox.pack(side='right')
        iter_spinbox.bind('<Return>', lambda e: self.update_step7_iterations(iter_var))
        iter_spinbox.bind('<FocusOut>', lambda e: self.update_step7_iterations(iter_var))
        self.param_widgets = {'iter_var': iter_var}

    def create_step8_controls(self):
        """Create controls for Step 8: Morphological Open"""
        skip_var = tk.BooleanVar(value=not self.params['step_8_morph_open']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_8_morph_open', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(self.params_frame, text="Number of Iterations:").pack(anchor='w', pady=(5, 2))
        iter_frame = ttk.Frame(self.params_frame)
        iter_frame.pack(fill='x', pady=2)
        iter_var = tk.IntVar(value=self.params['step_8_morph_open']['iterations'])
        ttk.Scale(iter_frame, from_=1, to=10, orient='horizontal', variable=iter_var,
                  command=lambda v: self.update_step8_iterations(iter_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        iter_spinbox = ttk.Spinbox(iter_frame, from_=1, to=10, textvariable=iter_var,
                                   width=6, command=lambda: self.update_step8_iterations(iter_var))
        iter_spinbox.pack(side='right')
        iter_spinbox.bind('<Return>', lambda e: self.update_step8_iterations(iter_var))
        iter_spinbox.bind('<FocusOut>', lambda e: self.update_step8_iterations(iter_var))
        self.param_widgets = {'iter_var': iter_var}

    def create_step9_controls(self):
        """Create controls for Step 9: Fill Holes"""
        skip_var = tk.BooleanVar(value=not self.params['step_9_fillholes']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_9_fillholes', skip_var)).pack(anchor='w', pady=5)

        ttk.Label(self.params_frame, text="Note: Fills small holes inside detected regions.",
                 wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=10)

    def create_step10_controls(self):
        """Create controls for Step 10: Labeling"""
        skip_var = tk.BooleanVar(value=not self.params['step_10_labeling']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                       variable=skip_var,
                       command=lambda: self.toggle_step_skip('step_10_labeling', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(self.params_frame, text="Minimum Size (pixels):").pack(anchor='w', pady=(5, 2))
        size_frame = ttk.Frame(self.params_frame)
        size_frame.pack(fill='x', pady=2)
        size_var = tk.IntVar(value=self.params['step_10_labeling']['min_size'])
        ttk.Scale(size_frame, from_=0, to=200, orient='horizontal', variable=size_var,
                  command=lambda v: self.update_step10_size(size_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        size_spinbox = ttk.Spinbox(size_frame, from_=0, to=200, textvariable=size_var,
                                   width=6, command=lambda: self.update_step10_size(size_var))
        size_spinbox.pack(side='right')
        size_spinbox.bind('<Return>', lambda e: self.update_step10_size(size_var))
        size_spinbox.bind('<FocusOut>', lambda e: self.update_step10_size(size_var))
        self.param_widgets = {'size_var': size_var}

    def create_step11_controls(self):
        """Create controls for Step 11: Final Result"""
        ttk.Label(self.params_frame, text="Minimum Area (pixels):").pack(anchor='w', pady=(5, 2))
        min_area_frame = ttk.Frame(self.params_frame)
        min_area_frame.pack(fill='x', pady=2)
        min_area_var = tk.IntVar(value=self.params['step_11_final']['min_area'])
        ttk.Scale(min_area_frame, from_=0, to=1500, orient='horizontal', variable=min_area_var,
                  command=lambda v: self.update_step11_min_area(min_area_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        min_spinbox = ttk.Spinbox(min_area_frame, from_=0, to=1500, textvariable=min_area_var,
                                  width=6, command=lambda: self.update_step11_min_area(min_area_var))
        min_spinbox.pack(side='right')
        min_spinbox.bind('<Return>', lambda e: self.update_step11_min_area(min_area_var))
        min_spinbox.bind('<FocusOut>', lambda e: self.update_step11_min_area(min_area_var))

        ttk.Label(self.params_frame, text="Maximum Area (pixels):").pack(anchor='w', pady=(10, 2))
        max_area_frame = ttk.Frame(self.params_frame)
        max_area_frame.pack(fill='x', pady=2)
        max_area_var = tk.IntVar(value=self.params['step_11_final']['max_area'])
        ttk.Scale(max_area_frame, from_=1000, to=50000, orient='horizontal', variable=max_area_var,
                  command=lambda v: self.update_step11_max_area(max_area_var)).pack(side='left', fill='x', expand=True, padx=(0, 5))
        max_spinbox = ttk.Spinbox(max_area_frame, from_=1000, to=50000, textvariable=max_area_var,
                                  width=6, command=lambda: self.update_step11_max_area(max_area_var))
        max_spinbox.pack(side='right')
        max_spinbox.bind('<Return>', lambda e: self.update_step11_max_area(max_area_var))
        max_spinbox.bind('<FocusOut>', lambda e: self.update_step11_max_area(max_area_var))
        ttk.Label(self.params_frame, text="Note: Segments are colored red.",
                 wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=10)
        self.param_widgets = {'min_area_var': min_area_var, 'max_area_var': max_area_var}

    def create_step12_controls(self):
        """Create controls for Step 12: Perspective Warp"""
        skip_var = tk.BooleanVar(value=not self.params['step_12_perspective']['enabled'])
        ttk.Checkbutton(self.params_frame, text="Skip this step",
                        variable=skip_var,
                        command=lambda: self.toggle_step_skip('step_12_perspective', skip_var)).pack(anchor='w', pady=5)

        ttk.Separator(self.params_frame, orient='horizontal').pack(fill='x', pady=10)
        ttk.Label(self.params_frame,
                  text="Perspective warp rectifies the ROI polygon to a top-down view.\n\n"
                       "The 4 extreme corners of your ROI polygon (top-left, top-right, "
                       "bottom-right, bottom-left) are used as the source quad and mapped "
                       "to a fixed 272 × 272 px output (representing the 272 mm × 272 mm "
                       "panel at 10 mm height).",
                  wraplength=280, font=('Arial', 8)).pack(anchor='w', pady=5)

    # Parameter update methods
    def toggle_step_skip(self, step_key, skip_var):
        """Toggle skip for any step"""
        self.params[step_key]['enabled'] = not skip_var.get()
        self.process_all_steps()
        self.display_current_step()

    def update_step2_rotate(self, rot_var):
        """Update step 2 rotation angle"""
        try:
            value = int(rot_var.get())
            value = max(-180, min(180, value))
            rot_var.set(value)
            self.params['step_2_rotate_flip']['rotate'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            rot_var.set(self.params['step_2_rotate_flip']['rotate'])

    def update_step2_flip(self, flip_h_var, flip_v_var):
        """Update step 2 flip flags"""
        if flip_h_var is not None:
            self.params['step_2_rotate_flip']['flip_h'] = flip_h_var.get()
        if flip_v_var is not None:
            self.params['step_2_rotate_flip']['flip_v'] = flip_v_var.get()
        self.process_all_steps()
        self.display_current_step()

    def update_step5_params(self, low_var, high_var):
        """Update step 5 threshold parameters"""
        try:
            if low_var:
                value = max(0, min(255, int(low_var.get())))
                low_var.set(value)
                self.params['step_5_markers']['low_threshold'] = value
            if high_var:
                value = max(0, min(255, int(high_var.get())))
                high_var.set(value)
                self.params['step_5_markers']['high_threshold'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            if low_var:
                low_var.set(self.params['step_5_markers']['low_threshold'])
            if high_var:
                high_var.set(self.params['step_5_markers']['high_threshold'])

    def update_step7_iterations(self, iter_var):
        """Update step 7 morphological close iterations"""
        try:
            value = max(1, min(10, int(iter_var.get())))
            iter_var.set(value)
            self.params['step_7_morph_close']['iterations'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            iter_var.set(self.params['step_7_morph_close']['iterations'])

    def update_step8_iterations(self, iter_var):
        """Update step 8 morphological open iterations"""
        try:
            value = max(1, min(10, int(iter_var.get())))
            iter_var.set(value)
            self.params['step_8_morph_open']['iterations'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            iter_var.set(self.params['step_8_morph_open']['iterations'])

    def update_step10_size(self, size_var):
        """Update step 10 minimum size"""
        try:
            value = max(0, min(200, int(size_var.get())))
            size_var.set(value)
            self.params['step_10_labeling']['min_size'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            size_var.set(self.params['step_10_labeling']['min_size'])

    def update_step11_min_area(self, area_var):
        """Update step 11 minimum area"""
        try:
            value = max(0, min(1500, int(area_var.get())))
            area_var.set(value)
            self.params['step_11_final']['min_area'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            area_var.set(self.params['step_11_final']['min_area'])

    def update_step11_max_area(self, area_var):
        """Update step 11 maximum area"""
        try:
            value = max(1000, min(50000, int(area_var.get())))
            area_var.set(value)
            self.params['step_11_final']['max_area'] = value
            self.process_all_steps()
            self.display_current_step()
        except ValueError:
            area_var.set(self.params['step_11_final']['max_area'])

    def tab2_previous_step(self):
        """Navigate to previous step"""
        if self.current_step > 0:
            self.current_step -= 1
            # Going back to Step 1 re-enters drawing mode (keep existing points)
            self.display_current_step()

    def tab2_next_step(self):
        """Navigate to next step"""
        # Moving from Step 1 → Step 2: ROI must be drawn and closed first
        if self.current_step == 1:
            if not self.step1_polygon_closed or len(self.step1_polygon_points) < 3:
                messagebox.showwarning("Warning",
                    "Please complete the ROI polygon first.\n"
                    "Click points on the image, then click near the first point to close.")
                return
            # Run full pipeline (steps 2-12) now that ROI is confirmed
            self.process_all_steps()
            self.current_step = 2
            self.display_current_step()
            self.tab2_info_label.config(
                text="ROI confirmed. Processing complete. Navigate through steps.")
            return

        if self.current_step < 12:
            self.current_step += 1
            self.display_current_step()

    def tab2_save_config(self):
        """Save current configuration to JSON.
           Defaults to the same folder as the auto-saved ROI JSON."""
        if self.tab2_roi_points is None:
            messagebox.showwarning("Warning",
                "No ROI defined yet. Please complete Step 1 ROI drawing first.")
            return

        # Default save directory = same as ROI auto-save (next to image)
        init_dir = getattr(self, '_roi_save_dir', None)
        if init_dir is None and self.tab2_image_path:
            init_dir = os.path.dirname(os.path.abspath(self.tab2_image_path))

        img_stem = Path(self.tab2_image_path).stem if self.tab2_image_path else "config"
        default_name = f"{img_stem}_config.json"

        save_path = filedialog.asksaveasfilename(
            title="Save Configuration",
            initialdir=init_dir,
            initialfile=default_name,
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if save_path:
            data = {
                "image_path": self.tab2_image_path,
                "polygon_points": self.tab2_roi_points.tolist(),
                "pipeline_params": self.params
            }

            with open(save_path, 'w') as f:
                json.dump(data, f, indent=4)

            # ── Also write a copy to configurations/conf.json so Batch tab
            #    can auto-load it on next startup ──────────────────────────
            try:
                conf_dir = Path(__file__).resolve().parent / 'configurations'
                conf_dir.mkdir(exist_ok=True)
                with open(conf_dir / 'conf.json', 'w') as _cf:
                    json.dump(data, _cf, indent=4)
            except Exception:
                pass  # never block the main save if this fails

            messagebox.showinfo("Success", f"Configuration saved to {save_path}")

    def tab2_load_config(self):
        """Load configuration from JSON.
        When polygon_points are present, pre-loads them into the Step 1 drawing
        state so that clicking 'Start Processing' will show the saved polygon
        in Step 1 rather than an empty canvas.
        """
        file_path = filedialog.askopenfilename(
            title="Load Configuration",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            with open(file_path, 'r') as f:
                data = json.load(f)

            # Load ROI points into BOTH the roi_points array AND the Step 1
            # polygon drawing state so Start Processing can restore them.
            if 'polygon_points' in data:
                pts = data['polygon_points']
                self.tab2_roi_points       = np.array(pts, dtype=np.int32)
                # Pre-fill Step 1 drawing state
                self.step1_polygon_points  = [tuple(p) for p in pts]
                self.step1_polygon_closed  = True
            else:
                # Config has no polygon — clear any previously loaded ROI
                self.tab2_roi_points      = None
                self.step1_polygon_points = []
                self.step1_polygon_closed = False

            # Load pipeline parameters
            if 'pipeline_params' in data:
                self.params = data['pipeline_params']

            messagebox.showinfo("Success", "Configuration loaded successfully")
            self.tab2_info_label.config(
                text=f"Configuration loaded "
                     f"({'ROI ready – click Start Processing to review in Step 1' if self.tab2_roi_points is not None else 'No ROI in file – draw ROI in Step 1'}).")

    # ==================== TAB 3: BATCH PROCESSING ====================

    def init_tab3(self):
        """Stub — the merged tab UI is built by _init_tab_batch_3d."""
        pass  # All GUI and state initialization moved to _init_tab_batch_3d

    def _tab3_select_output_folder(self):
        """Select output folder."""
        folder_path = filedialog.askdirectory(title="Select Output Folder")
        if folder_path:
            self._tab3_output_var.set(folder_path)

    def tab3_select_folder(self):
        """Select input folder for batch processing"""
        folder_path = filedialog.askdirectory(title="Select Input Folder")

        if folder_path:
            self.tab3_input_folder = folder_path
            self._tab3_folder_var.set(folder_path)
            self.tab3_log(f"Input folder selected: {folder_path}")
            self.tab3_check_ready()

    def tab3_load_config(self):
        """Load configuration file for batch processing"""
        file_path = filedialog.askopenfilename(
            title="Load Configuration",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'r') as f:
                    self.tab3_config_data = json.load(f)

                # Verify it has required keys
                if 'polygon_points' not in self.tab3_config_data or 'pipeline_params' not in self.tab3_config_data:
                    messagebox.showerror("Error",
                                          "Invalid configuration file. Must contain polygon_points and pipeline_params.")
                    self.tab3_config_data = None
                    return

                self.tab3_log(f"Configuration loaded: {file_path}")
                self.tab3_check_ready()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load configuration: {str(e)}")
                self.tab3_config_data = None

    def tab3_load_roi(self):
        """Load ROI coordinates from JSON file"""
        file_path = filedialog.askopenfilename(
            title="Select ROI Coordinates File",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if file_path:
            try:
                with open(file_path, 'r') as f:
                    data = json.load(f)

                # Can load from either pure ROI file or config file
                if 'polygon_points' in data:
                    self.tab3_roi_points = np.array(data['polygon_points'], dtype=np.int32)
        
                    self.tab3_log(f"ROI coordinates loaded: {file_path} ({len(self.tab3_roi_points)} points)")
                    self.tab3_check_ready()
                else:
                    messagebox.showerror("Error", "Invalid file. Must contain 'polygon_points'.")
                    self.tab3_roi_points = None
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load ROI: {str(e)}")
                self.tab3_roi_points = None

    def tab3_check_ready(self):
        """Check if ready to start batch processing"""
        # Need: input folder + ROI points
        if self.tab3_input_folder and self.tab3_roi_points is not None:
            self._batch3d_start_btn.config(state='normal')
        elif self.tab3_input_folder:
            # Allow start even without explicit ROI; pipeline will use defaults
            self._batch3d_start_btn.config(state='normal')
        else:
            self._batch3d_start_btn.config(state='disabled')

    def tab3_log(self, message):
        """Add message to log (GUI) and mirror it to the terminal in real time.

        This is the single choke point through which ALL batch-processing,
        report-generation and 3D Slicer status/progress/warning/completion
        messages flow (including the ones streamed live from the Slicer
        subprocess in `_batch3d_run_viz`). Mirroring here guarantees that
        every message the user sees in the GUI log is also printed to the
        terminal/console, with no changes required at any call site and no
        change to the existing GUI behaviour.
        """
        # ── Existing GUI behaviour (unchanged) ───────────────────────────
        self.tab3_log_text.insert(tk.END, message + '\n')
        self.tab3_log_text.see(tk.END)
        self.tab3_log_text.update()

        # ── New: real-time terminal/console mirror ───────────────────────
        try:
            timestamp = datetime.now().strftime('%H:%M:%S')
            print(f"[{timestamp}] {message}", flush=True)
        except Exception:
            # Never let a terminal/console print failure break the GUI log.
            try:
                print(message, flush=True)
            except Exception:
                pass

    def tab3_start_batch_processing(self):
        """Start batch processing of images — runs in background thread."""
        if self.tab3_processing:
            return

        self.tab3_processing = True
        self._batch3d_start_btn.config(state='disabled')

        # Collect image files
        image_extensions = ['.png', '.jpg', '.jpeg']
        image_files = [
            os.path.join(self.tab3_input_folder, f)
            for f in os.listdir(self.tab3_input_folder)
            if os.path.isfile(os.path.join(self.tab3_input_folder, f))
            and any(f.lower().endswith(ext) for ext in image_extensions)
        ]

        if not image_files:
            messagebox.showwarning("Warning", "No image files found in selected folder")
            self.tab3_processing = False
            self._batch3d_start_btn.config(state='normal')
            return

        self.tab3_log(f"\nFound {len(image_files)} image(s) to process")

        # Output folders
        output_base = self._tab3_output_var.get()
        if not output_base or output_base == 'Not selected':
            output_base = self.tab3_input_folder
        defects_path     = os.path.join(output_base, "defects")
        non_defects_path = os.path.join(output_base, "non_defects")
        os.makedirs(defects_path, exist_ok=True)
        os.makedirs(non_defects_path, exist_ok=True)

        self.tab3_log(f"Output folders:")
        self.tab3_log(f"  - {defects_path}")
        self.tab3_log(f"  - {non_defects_path}")

        # Snapshot config & ROI so the background thread doesn't touch self
        roi_points  = (self.tab3_roi_points.tolist()
                       if self.tab3_roi_points is not None else None)
        config_data = self.tab3_config_data

        # ── Run in background thread so GUI stays responsive ─────────────
        def _run():
            total         = len(image_files)
            defect_count  = 0
            non_defect    = 0
            # Determine worker count: min(CPU cores, 4) to stay memory-safe
            import multiprocessing as _mp
            workers = max(1, min(_mp.cpu_count(), 32))
            self.root.after(0, lambda: self.tab3_log(
                f"Starting parallel processing with {workers} worker(s)…"))

            # Use ThreadPoolExecutor — safe with Tkinter/OpenCV on all OSes
            # (ProcessPoolExecutor requires picklable args; threads share memory
            #  which is fine since each image write is to a different file)
            from concurrent.futures import ThreadPoolExecutor, as_completed
            futures = {}
            UPDATE_EVERY = max(1, total // 20)   # update UI every ~5 %

            with ThreadPoolExecutor(max_workers=workers) as executor:
                for img_path in image_files:
                    fut = executor.submit(
                        self.tab3_process_single_image,
                        img_path, defects_path, non_defects_path, img_path
                    )
                    futures[fut] = img_path

                for done_idx, fut in enumerate(as_completed(futures), start=1):
                    img_path = futures[fut]
                    fname    = os.path.basename(img_path)
                    try:
                        num_seg = fut.result()
                        if num_seg >= 1:
                            defect_count += 1
                            msg = f"✓ {fname} → defects ({num_seg} segments)"
                        else:
                            non_defect += 1
                            msg = f"✓ {fname} → non_defects"
                    except Exception as exc:
                        msg = f"✗ {fname} → ERROR: {exc}"

                    # Free result memory immediately
                    del fut
                    gc.collect()

                    # Throttle UI updates to avoid flooding the GUI thread
                    if done_idx % UPDATE_EVERY == 0 or done_idx == total:
                        progress = (done_idx / total) * 100
                        _msg = msg  # capture for lambda
                        _di  = done_idx
                        def _update(p=progress, m=_msg, di=_di):
                            self._batch3d_progress['value'] = p
                            self._batch3d_status_lbl.config(
                                text=f"Processing {di}/{total}…")
                            self.tab3_log(m)
                        self.root.after(0, _update)
                    else:
                        _msg = msg
                        self.root.after(0, lambda m=_msg: self.tab3_log(m))

            # ── Done ─────────────────────────────────────────────────────
            def _finish():
                self._batch3d_progress['value'] = 100
                self._batch3d_status_lbl.config(text="Processing Complete!")
                self.tab3_log(f"\n=== Processing Complete ===")
                self.tab3_log(f"Total: {total}  |  Defects: {defect_count}  |  Non-defects: {non_defect}")
                self.tab3_processing = False
                self._batch3d_start_btn.config(state='normal')
                messagebox.showinfo("Complete",
                    f"Batch processing complete!\n\n"
                    f"Processed : {total} images\n"
                    f"Defects   : {defect_count}\n"
                    f"Non-defects: {non_defect}")
            self.root.after(0, _finish)

        threading.Thread(target=_run, daemon=True).start()

    def tab3_process_single_image(self, image_path, defects_folder, non_defects_folder, filename):
        """Process a single image and save to appropriate folder"""
        # Load image
        cv_image = cv2.imread(image_path)
        if cv_image is None:
            raise ValueError("Failed to load image")

        # Get ROI points - use separately loaded ROI if available
        if self.tab3_roi_points is not None:
            roi_points = self.tab3_roi_points
        elif self.tab3_config_data and 'polygon_points' in self.tab3_config_data:
            roi_points = np.array(self.tab3_config_data['polygon_points'], dtype=np.int32)
        else:
            # No ROI defined – use full image as ROI
            h, w = cv_image.shape[:2]
            roi_points = np.array([[0, 0], [w-1, 0], [w-1, h-1], [0, h-1]], dtype=np.int32)

        # Get parameters from config, or fall back to built-in defaults
        # Keys match exactly those used by the Single Image (Tab 2) pipeline
        if self.tab3_config_data and 'pipeline_params' in self.tab3_config_data:
            params = self.tab3_config_data['pipeline_params']
        else:
            params = {
                'step_2_rotate_flip': {'enabled': True, 'rotate': 0, 'flip_h': False, 'flip_v': False},
                'step_3_grayscale': {'enabled': True},
                'step_4_sobel': {'enabled': False},
                'step_5_markers': {'enabled': True, 'low_threshold': 90, 'high_threshold': 220},
                'step_6_watershed': {'enabled': False},
                'step_7_morph_close': {'enabled': False, 'iterations': 1},
                'step_8_morph_open': {'enabled': False, 'iterations': 1},
                'step_9_fillholes': {'enabled': False},
                'step_10_labeling': {'enabled': True, 'min_size': 50},
                'step_11_final': {'min_area': 50, 'max_area': 10000},
            }
        # Step 1: Apply ROI mask
        mask = np.zeros(cv_image.shape[:2], dtype=np.uint8)
        cv2.fillPoly(mask, [roi_points], 255)
        masked_image = cv_image.copy()
        masked_image[mask == 0] = [0, 0, 0]

        # Step 2: Rotate / Flip
        rotated_bgr = masked_image.copy()
        if params.get('step_2_rotate_flip', {}).get('enabled', False):
            angle = params['step_2_rotate_flip'].get('rotate', 0)
            if angle != 0:
                h, w = rotated_bgr.shape[:2]
                M_rot = cv2.getRotationMatrix2D((w // 2, h // 2), -angle, 1.0)
                rotated_bgr = cv2.warpAffine(rotated_bgr, M_rot, (w, h),
                                             borderMode=cv2.BORDER_CONSTANT,
                                             borderValue=(0, 0, 0))
            if params['step_2_rotate_flip'].get('flip_h', False):
                rotated_bgr = cv2.flip(rotated_bgr, 1)
            if params['step_2_rotate_flip'].get('flip_v', False):
                rotated_bgr = cv2.flip(rotated_bgr, 0)

        # Step 3: Grayscale
        gray = cv2.cvtColor(rotated_bgr, cv2.COLOR_BGR2GRAY)

        # Step 4: Sobel gradient
        if params.get('step_4_sobel', {}).get('enabled', True):
            sobel_x = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
            sobel_y = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
            elevation_map = np.hypot(sobel_x, sobel_y)
            elevation_display = cv2.normalize(elevation_map, None, 0, 255,
                                              cv2.NORM_MINMAX).astype(np.uint8)
        else:
            elevation_display = gray

        # Step 5: Histogram markers
        #high_thresh = 150  # default fallback
        #foreground_mask = gray > high_thresh  # default fallback
        if params.get('step_5_markers', {}).get('enabled', True):
            low_thresh  = params['step_5_markers'].get('low_threshold', 30)
            high_thresh = params['step_5_markers'].get('high_threshold', 150)

            inv = 255 - gray
            background = cv2.GaussianBlur(inv, (0, 0), sigmaX=50, sigmaY=50)
            gbImg = cv2.subtract(inv, background)
            _, gb = cv2.threshold(gbImg, 10, 255, cv2.THRESH_BINARY)

            markers = np.zeros_like(gray, dtype=np.int32)
            background_mask = gray < low_thresh
            markers[background_mask] = gb[background_mask]
            foreground_mask = gray > high_thresh
            markers[foreground_mask] = gray[foreground_mask].astype(np.int32) + 1
        else:
            markers = np.zeros_like(gray, dtype=np.int32)
            markers[gray > 0] = gray[gray > 0].astype(np.int32) + 1

        # Step 6: Watershed
        if params.get('step_6_watershed', {}).get('enabled', True):
            elevation_map_color = cv2.cvtColor(elevation_display, cv2.COLOR_GRAY2BGR)
            markers_watershed = markers.copy()
            segmentation = cv2.watershed(elevation_map_color, markers_watershed)
        else:
            segmentation = markers.copy()

        # Convert segmentation to binary.
        # FIX: watershed labels can span the entire image (no true background seeds),
        # so thresholding on label values captures everything → one giant blob > max_area.
        # Instead, use the foreground_mask (bright defect pixels) directly, then let
        # morphological steps clean / separate the regions.
        if params.get('step_5_markers', {}).get('enabled', True):
            # binary_seg = (segmentation > high_thresh).astype(np.uint8)
            high_thresh = params['step_5_markers'].get('high_threshold', 150)
            binary_seg = (segmentation > high_thresh).astype(np.uint8)
        else:
            binary_seg = (segmentation > 0).astype(np.uint8)

        # Step 7: Morphological Close
        if params.get('step_7_morph_close', {}).get('enabled', True):
            iterations = params['step_7_morph_close'].get('iterations', 1)
            kernel = np.ones((3, 3), np.uint8)
            morph_close = cv2.morphologyEx(binary_seg, cv2.MORPH_CLOSE, kernel,
                                           iterations=iterations)
        else:
            morph_close = binary_seg.copy()

        # Step 8: Morphological Open
        if params.get('step_8_morph_open', {}).get('enabled', True):
            iterations = params['step_8_morph_open'].get('iterations', 1)
            kernel = np.ones((3, 3), np.uint8)
            morph_open = cv2.morphologyEx(morph_close, cv2.MORPH_OPEN, kernel,
                                          iterations=iterations)
        else:
            morph_open = morph_close.copy()

        # Step 9: Fill holes
        if params.get('step_9_fillholes', {}).get('enabled', True):
            filled = ndi.binary_fill_holes(morph_open).astype(np.uint8) * 255
        else:
            filled = morph_open * 255

        # Step 10: Label components
        if params.get('step_10_labeling', {}).get('enabled', True):
            labeled_array, num_features = ndi.label(filled)
            min_size = params['step_10_labeling'].get('min_size', 50)
            sizes = np.bincount(labeled_array.ravel())
            mask_sizes = sizes > min_size
            mask_sizes[0] = 0
            cleaned_labels = mask_sizes[labeled_array]
            labeled_cleaned, num_cleaned = ndi.label(cleaned_labels)
        else:
            labeled_cleaned, num_cleaned = ndi.label(filled)

        # Step 11: Color valid segments on original image
        result = cv_image.copy()
        segment_color = (0, 0, 255)  # Red in BGR
        min_area = params.get('step_11_final', {}).get('min_area', 50)
        max_area = params.get('step_11_final', {}).get('max_area', 10000)

        valid_segments = 0
        for region_label in range(1, num_cleaned + 1):
            region_mask = (labeled_cleaned == region_label).astype(np.uint8) * 255
            contours, _ = cv2.findContours(region_mask, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
            if contours:
                for contour in contours:
                    area = cv2.contourArea(contour)
                    if min_area <= area <= max_area:
                        valid_segments += 1
                        cv2.drawContours(result, [contour], -1, segment_color, -1)

        # Save to appropriate folder
        filename = os.path.basename(filename)

        # Step 12: Perspective Warp — apply before saving
        warp_enabled = params.get('step_12_perspective', {}).get('enabled', True)
        if warp_enabled and roi_points is not None and len(roi_points) >= 4:
            try:
                roi_pts = roi_points.reshape(-1, 2).astype(np.float32)
                s = roi_pts.sum(axis=1)
                d = np.diff(roi_pts, axis=1).ravel()
                tl = roi_pts[np.argmin(s)]
                br = roi_pts[np.argmax(s)]
                tr = roi_pts[np.argmin(d)]
                bl = roi_pts[np.argmax(d)]
                src_quad = np.array([tl, tr, br, bl], dtype=np.float32)
                out_w, out_h = 272*8, 272*8
                dst_quad = np.array([[0, 0], [out_w - 1, 0],
                                     [out_w - 1, out_h - 1], [0, out_h - 1]], dtype=np.float32)
                M = cv2.getPerspectiveTransform(src_quad, dst_quad)
                result = cv2.warpPerspective(result, M, (out_w, out_h),
                                             flags=cv2.INTER_LINEAR,
                                             borderMode=cv2.BORDER_CONSTANT,
                                             borderValue=(0, 0, 0))
            except Exception:
                pass  # if warp fails, fall back to un-warped result

        if valid_segments >= 1:
            output_path = os.path.join(defects_folder, filename)
        else:
            output_path = os.path.join(non_defects_folder, filename)

        cv2.imwrite(output_path, result)

        # ── Explicit memory cleanup — free all intermediate arrays ───────
        del cv_image, masked_image, rotated_bgr, gray, result
        try:
            del elevation_display, elevation_map
        except Exception:
            pass
        try:
            del markers, watershed_result, binary
        except Exception:
            pass
        gc.collect()

        return valid_segments

    # ==================== REPORTS ====================

    def _init_tab_reports(self):
        """Stub — Reports state initialized in _init_tab_batch_3d.
        The Reports UI itself is built on-demand inside a popup window by
        _open_reports_popup(), mirroring the existing independent 3D
        Visualise dialog, so no new tab/page is added to the application."""
        pass

    def _open_reports_popup(self):
        """Reports — opens a dedicated dialog (mirrors the independent 3D
        Visualise popup pattern). Contains the exact Reports functionality
        ported from the reference implementation: select Predicted Results
        and Ground Truth folders, Generate Report (binary classification
        evaluation), and Save Excel (two-sheet workbook export)."""

        popup = tk.Toplevel(self.root)
        popup.title("Reports — Binary Classification Evaluation")
        popup.resizable(False, False)
        popup.configure(bg=self.BG)
        popup.grab_set()

        # Centre over main window
        self.root.update_idletasks()
        rx = self.root.winfo_x() + (self.root.winfo_width()  - 760) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - 520) // 2
        popup.geometry(f"760x520+{rx}+{ry}")

        # ── Scrollable container (same pattern as final_app2 Reports tab) ──
        outer = tk.Frame(popup, bg=self.BG)
        outer.pack(fill='both', expand=True)

        _canvas = tk.Canvas(outer, bg=self.BG, highlightthickness=0)
        _vsb = ttk.Scrollbar(outer, orient='vertical', command=_canvas.yview)
        _canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side='right', fill='y')
        _canvas.pack(side='left', fill='both', expand=True)

        card = tk.Frame(_canvas, bg=self.PANEL, bd=0, relief='flat')
        _card_id = _canvas.create_window((0, 0), window=card, anchor='n')

        def _on_card_resize(e):
            _canvas.configure(scrollregion=_canvas.bbox('all'))

        def _on_canvas_resize(e):
            w = min(720, e.width)
            _canvas.itemconfig(_card_id, width=w)
            _canvas.coords(_card_id, e.width // 2, 0)
            card.configure(width=w)

        card.bind('<Configure>', _on_card_resize)
        _canvas.bind('<Configure>', _on_canvas_resize)

        def _on_mw(e):
            _canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        _canvas.bind_all('<MouseWheel>', _on_mw)

        # ── Title ────────────────────────────────────────────────────────
        tk.Label(card, text='Reports', bg=self.PANEL, fg=self.HIGHLIGHT,
                 font=('Segoe UI', 18, 'bold')).pack(pady=(24, 2))
        tk.Label(card, text='Binary Classification Evaluation — Compare Predictions vs Ground Truth',
                 bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9)).pack(pady=(0, 6))
        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=30, pady=10)

        # ── Helper: browse row ───────────────────────────────────────────
        def _browse_row(parent, label, var, pick_fn):
            tk.Label(parent, text=label, bg=self.PANEL, fg=self.FG2,
                     font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=30)
            row = tk.Frame(parent, bg=self.PANEL)
            row.pack(fill='x', padx=30, pady=(2, 10))
            tk.Entry(row, textvariable=var, bg=self.ENTRY_BG, fg=self.FG2,
                     relief='flat', font=('Segoe UI', 9), state='readonly',
                     readonlybackground=self.ENTRY_BG, bd=0
                     ).pack(side='left', fill='x', expand=True, ipady=6, padx=(0, 8))
            tk.Button(row, text='Browse', command=pick_fn,
                      bg=self.BTN_BG, fg=self.FG, relief='flat',
                      font=('Segoe UI', 9, 'bold'), padx=10, pady=3,
                      activebackground=self.HIGHLIGHT, cursor='hand2').pack(side='right')

        # ── Folder inputs ────────────────────────────────────────────────
        self._rpt_pred_var = tk.StringVar(value='No folder selected')
        self._rpt_gt_var   = tk.StringVar(value='No folder selected')

        def _pick_pred():
            p = filedialog.askdirectory(
                title='Select Predicted Results Folder (contains Indicator / Non Indicator subfolders)',
                parent=popup)
            if p:
                self._rpt_pred_var.set(p)

        def _pick_gt():
            p = filedialog.askdirectory(
                title='Select Ground Truth Folder (contains Indicator / Non Indicator subfolders)',
                parent=popup)
            if p:
                self._rpt_gt_var.set(p)

        _browse_row(card, 'Predicted Results Folder', self._rpt_pred_var, _pick_pred)
        _browse_row(card, 'Ground Truth Folder',      self._rpt_gt_var,   _pick_gt)

        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=30, pady=6)

        # ── Action buttons ───────────────────────────────────────────────
        btn_row = tk.Frame(card, bg=self.PANEL)
        btn_row.pack(pady=(8, 4))

        self._rpt_generate_btn = tk.Button(
            btn_row, text='  ▶  Generate Report',
            bg=self.HIGHLIGHT, fg='#ffffff', relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=22, pady=8,
            activebackground=self.BTN_ACT, cursor='hand2',
            command=self._rpt_generate)
        self._rpt_generate_btn.pack(side='left', padx=(0, 10))

        self._rpt_save_btn = tk.Button(
            btn_row, text='  💾  Save Excel',
            bg=self.BTN_BG, fg=self.FG, relief='flat',
            font=('Segoe UI', 11, 'bold'), padx=22, pady=8,
            activebackground=self.HIGHLIGHT, cursor='hand2',
            state='disabled',
            command=self._rpt_save_excel)
        self._rpt_save_btn.pack(side='left')

        tk.Frame(card, bg=self.ACCENT, height=1).pack(fill='x', padx=30, pady=(12, 6))

        # ── Status / log label ───────────────────────────────────────────
        self._rpt_status_var = tk.StringVar(value='Select both folders and click Generate Report.')
        tk.Label(card, textvariable=self._rpt_status_var,
                 bg=self.PANEL, fg=self.FG2,
                 font=('Segoe UI', 9), wraplength=620, justify='left'
                 ).pack(anchor='w', padx=30, pady=(4, 24))

        close_row = tk.Frame(card, bg=self.PANEL)
        close_row.pack(pady=(0, 18))
        tk.Button(close_row, text='Close',
                  command=popup.destroy,
                  bg=self.BTN_BG, fg=self.FG, relief='flat',
                  font=('Segoe UI', 10, 'bold'), padx=14, pady=7,
                  activebackground=self.HIGHLIGHT, cursor='hand2').pack()

    # ── Reports: internal helpers ────────────────────────────────────────

    def _rpt_generate(self):
        """Collect folders, match images, compute metrics, update UI."""
        pred_folder = self._rpt_pred_var.get().strip()
        gt_folder   = self._rpt_gt_var.get().strip()

        if pred_folder == 'No folder selected' or not os.path.isdir(pred_folder):
            messagebox.showwarning('Reports', 'Please select a valid Predicted Results folder.')
            return
        if gt_folder == 'No folder selected' or not os.path.isdir(gt_folder):
            messagebox.showwarning('Reports', 'Please select a valid Ground Truth folder.')
            return

        self._rpt_status_var.set('Processing…')
        self._rpt_generate_btn.config(state='disabled')
        self.root.update_idletasks()

        try:
            results, metrics, warnings = self._rpt_evaluate(pred_folder, gt_folder)
        except Exception as exc:
            messagebox.showerror('Reports Error', f'Evaluation failed:\n{exc}')
            self._rpt_generate_btn.config(state='normal')
            self._rpt_status_var.set('Error during evaluation. Check folder structure.')
            return

        # Store for later saving
        self._rpt_results_data = results
        self._rpt_metrics_data = metrics
        self._rpt_pred_folder  = pred_folder
        self._rpt_gt_folder    = gt_folder

        # Update metric tiles
        total = len(results)
        warn_txt = f'  ⚠ {len(warnings)} warning(s): {"; ".join(warnings[:3])}{"…" if len(warnings) > 3 else ""}' if warnings else ''
        self._rpt_status_var.set(
            f'✔  Evaluation complete — {total} image(s) processed. '
            f'TP={metrics["tp"]}  TN={metrics["tn"]}  FP={metrics["fp"]}  FN={metrics["fn"]}.{warn_txt}  '
            f'Click "Save Excel" to export the report.')

        self._rpt_save_btn.config(state='normal')
        self._rpt_generate_btn.config(state='normal')

    def _rpt_evaluate(self, pred_folder, gt_folder):
        """
        Match images by filename, compute TP/TN/FP/FN.
        Returns (results_list, metrics_dict, warnings_list).
        """
        DEFECT     = 'defects'
        NON_DEFECT = 'non_defects'

        def _load_folder(root_path):
            """Return {filename_lower: category_string} from defects/ non_defects/ subfolders."""
            mapping = {}
            for sub in os.listdir(root_path):
                sub_path = os.path.join(root_path, sub)
                if not os.path.isdir(sub_path):
                    continue
                sub_lower = sub.lower().strip()
                # Normalise: accept 'defect', 'defects', 'non_defect', 'non_defects', etc.
                if sub_lower in ('defects', 'defect'):
                    category = DEFECT
                elif sub_lower in ('non_defects', 'non_defect'):
                    category = NON_DEFECT
                else:
                    continue  # skip unknown subfolders
                for fname in os.listdir(sub_path):
                    if fname.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.tif')):
                        mapping[fname.lower()] = category
            return mapping

        pred_map = _load_folder(pred_folder)
        gt_map   = _load_folder(gt_folder)

        all_filenames = set(pred_map.keys()) | set(gt_map.keys())

        results  = []
        warnings = []
        tp = tn = fp = fn = 0

        for fname in sorted(all_filenames):
            gt_cat   = gt_map.get(fname)
            pred_cat = pred_map.get(fname)

            # ── Edge cases ───────────────────────────────────────────────
            if gt_cat is None:
                # In prediction but not in GT → note separately
                results.append({
                    'Image Name':   fname,
                    'Ground Truth': 'N/A (missing)',
                    'Predicted':    pred_cat,
                    'Result':       'No GT',
                })
                warnings.append(f'{fname}: exists in Prediction but not in Ground Truth')
                continue

            if pred_cat is None:
                # In GT but not in Prediction → treat as FN
                results.append({
                    'Image Name':   fname,
                    'Ground Truth': gt_cat,
                    'Predicted':    'N/A (missing)',
                    'Result':       'FN (missing prediction)',
                })
                fn += 1
                warnings.append(f'{fname}: exists in Ground Truth but not in Prediction (counted as FN)')
                continue

            # ── Normal classification ────────────────────────────────────
            if gt_cat == DEFECT and pred_cat == DEFECT:
                label = 'TP'; tp += 1
            elif gt_cat == NON_DEFECT and pred_cat == NON_DEFECT:
                label = 'TN'; tn += 1
            elif gt_cat == NON_DEFECT and pred_cat == DEFECT:
                label = 'FP'; fp += 1
            else:  # gt==defect, pred==non_defect
                label = 'FN'; fn += 1

            results.append({
                'Image Name':   fname,
                'Ground Truth': gt_cat,
                'Predicted':    pred_cat,
                'Result':       label,
            })

        # ── Metrics ──────────────────────────────────────────────────────
        total_classified = tp + tn + fp + fn
        accuracy  = (tp + tn) / total_classified if total_classified > 0 else 0.0
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall    = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1        = (2 * precision * recall / (precision + recall)
                     if (precision + recall) > 0 else 0.0)

        metrics = {
            'accuracy':  round(accuracy,  4),
            'precision': round(precision, 4),
            'recall':    round(recall,    4),
            'f1':        round(f1,        4),
            'tp': tp, 'tn': tn, 'fp': fp, 'fn': fn,
        }
        return results, metrics, warnings

    def _rpt_save_excel(self):
        """Save two-sheet Excel report to user-chosen path."""
        if not self._rpt_results_data or not self._rpt_metrics_data:
            messagebox.showwarning('Reports', 'No report data. Please generate a report first.')
            return

        save_path = filedialog.asksaveasfilename(
            title='Save Excel Report',
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            initialfile='classification_report.xlsx',
        )
        if not save_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                'Missing Library',
                'openpyxl is required to save Excel files.\n\n'
                'Install it with:  pip install openpyxl')
            return

        wb = openpyxl.Workbook()

        # ── Colour palette for Excel ─────────────────────────────────────
        HDR_FILL   = PatternFill('solid', fgColor='2D7DD2')
        HDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
        TP_FILL    = PatternFill('solid', fgColor='D1ECF1')
        TN_FILL    = PatternFill('solid', fgColor='D1ECF1')
        FP_FILL    = PatternFill('solid', fgColor='FFF3CD')
        FN_FILL    = PatternFill('solid', fgColor='FFF3CD')
        NOGT_FILL  = PatternFill('solid', fgColor='E2E3E5')
        THIN       = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        CENTER     = Alignment(horizontal='center', vertical='center')
        LEFT       = Alignment(horizontal='left',   vertical='center')

        def _hdr_cell(ws, row, col, value):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = HDR_FONT
            c.fill      = HDR_FILL
            c.alignment = CENTER
            c.border    = THIN
            return c

        # ── Sheet 1 : Detailed image-wise report ─────────────────────────
        ws1 = wb.active
        ws1.title = 'Image-wise Report'

        headers1 = ['Image Name', 'Ground Truth', 'Predicted', 'Result']
        col_widths1 = [32, 18, 18, 26]
        for col, (h, w) in enumerate(zip(headers1, col_widths1), start=1):
            _hdr_cell(ws1, 1, col, h)
            ws1.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w

        result_fill_map = {
            'TP': TP_FILL, 'TN': TN_FILL,
            'FP': FP_FILL, 'FN': FN_FILL,
        }

        for r_idx, row_data in enumerate(self._rpt_results_data, start=2):
            result_val = row_data['Result']
            row_fill   = result_fill_map.get(result_val[:2], NOGT_FILL)

            for col, key in enumerate(['Image Name', 'Ground Truth', 'Predicted', 'Result'],
                                       start=1):
                c = ws1.cell(row=r_idx, column=col, value=row_data[key])
                c.fill      = row_fill
                c.border    = THIN
                c.alignment = LEFT if col == 1 else CENTER
            ws1.row_dimensions[r_idx].height = 16

        ws1.freeze_panes = 'A2'

        # ── Sheet 2 : Overall metrics summary ────────────────────────────
        ws2 = wb.create_sheet(title='Overall Metrics')

        headers2 = ['Dataset / Path', 'Accuracy', 'Precision', 'Recall',
                    'F1 Score', 'TP', 'TN', 'FP', 'FN']
        col_widths2 = [48, 12, 12, 12, 12, 8, 8, 8, 8]
        for col, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
            _hdr_cell(ws2, 1, col, h)
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w

        m = self._rpt_metrics_data
        dataset_label = (f'Pred: {self._rpt_pred_folder}  |  GT: {self._rpt_gt_folder}')
        row_vals = [
            dataset_label,
            m['accuracy'], m['precision'], m['recall'], m['f1'],
            m['tp'], m['tn'], m['fp'], m['fn'],
        ]
        for col, val in enumerate(row_vals, start=1):
            c = ws2.cell(row=2, column=col, value=val)
            c.border    = THIN
            c.alignment = LEFT if col == 1 else CENTER
        ws2.row_dimensions[2].height = 18
        ws2.freeze_panes = 'A2'

        # ── Save ─────────────────────────────────────────────────────────
        try:
            wb.save(save_path)
            messagebox.showinfo('Reports', f'Excel report saved successfully:\n{save_path}')
            self._rpt_status_var.set(f'✔  Report saved to: {save_path}')
            # Keep the main tab's "latest report" hyperlink in sync (Change 3).
            self._batch3d_set_report_link(save_path)
        except Exception as exc:
            messagebox.showerror('Save Error', f'Could not save file:\n{exc}')


class LoginWindow:
    """Login screen shown before the main app."""

    USERNAME = 'admin'
    PASSWORD = 'admin'

    def __init__(self, root):
        self.root = root
        #self.root.title("Indicator Detection – Login")
        self.root.geometry("420x520")
        self.root.resizable(False, False)

        BG      = '#1e2328'
        PANEL   = '#252b33'
        ACCENT  = '#2e3a4a'
        BLUE    = '#2d7dd2'
        FG      = '#dde3ea'
        FG2     = '#7f8c9a'
        ENTRY   = '#181d22'

        self.root.configure(bg=BG)

        # ── Card ──────────────────────────────────────────────────────
        card = tk.Frame(self.root, bg=PANEL, bd=0)
        card.place(relx=0.5, rely=0.5, anchor='center', width=340, height=480)

        
        # Username
        tk.Label(card, text='Username', bg=PANEL, fg=FG2,
                 font=('Segoe UI', 9), anchor='w').pack(fill='x', padx=36)
        self.user_var = tk.StringVar()
        user_entry = tk.Entry(card, textvariable=self.user_var,
                              bg=ENTRY, fg=FG, relief='flat',
                              font=('Segoe UI', 11), insertbackground=FG, bd=0)
        user_entry.pack(fill='x', padx=36, ipady=8, pady=(3, 14))

        # Password
        tk.Label(card, text='Password', bg=PANEL, fg=FG2,
                 font=('Segoe UI', 9), anchor='w').pack(fill='x', padx=36)
        self.pass_var = tk.StringVar()
        self.pass_entry = tk.Entry(card, textvariable=self.pass_var,
                                   show='●', bg=ENTRY, fg=FG, relief='flat',
                                   font=('Segoe UI', 11), insertbackground=FG, bd=0)
        self.pass_entry.pack(fill='x', padx=36, ipady=8, pady=(3, 6))

        # Error label
        self.err_label = tk.Label(card, text='', bg=PANEL, fg='#e05c5c',
                                  font=('Segoe UI', 9))
        self.err_label.pack(pady=(0, 12))

        # Login button
        login_btn = tk.Button(card, text='Login',
                              command=self._try_login,
                              bg=BLUE, fg='#ffffff', relief='flat',
                              font=('Segoe UI', 11, 'bold'), padx=20, pady=8,
                              activebackground='#1a5fa8', activeforeground='#ffffff',
                              cursor='hand2')
        login_btn.pack(padx=36, fill='x')

        # Skip / continue as guest
        skip_btn = tk.Button(card, text='Continue without login \u2192',
                             command=self._skip_login,
                             bg=PANEL, fg=FG2, relief='flat',
                             font=('Segoe UI', 9), pady=4,
                             activebackground=PANEL, activeforeground=FG,
                             cursor='hand2', bd=0)
        skip_btn.pack(pady=(8, 0))

        # ── Bindings ─────────────────────────────────────────────────
        user_entry.bind('<Return>', lambda e: self.pass_entry.focus())
        self.pass_entry.bind('<Return>', lambda e: self._try_login())

        user_entry.focus()

    def _try_login(self):
        u = self.user_var.get().strip()
        p = self.pass_var.get().strip()
        if u == self.USERNAME and p == self.PASSWORD:
            self.root.destroy()
            _launch_main()
        else:
            self.err_label.config(text='Invalid username or password.')
            self.pass_var.set('')

    def _skip_login(self):
        self.root.destroy()
        _launch_main()


def _launch_main():
    root = tk.Tk()
    app = ImageAnalyzerApp(root)
    root.mainloop()


def _launch_login():
    root = tk.Tk()
    LoginWindow(root)
    root.mainloop()


class SplashScreen:
    """Animated splash / intro screen shown before the login window."""

    def __init__(self, root):
        self.root = root
        #self.root.title("Indicator Detection System")
        self.root.geometry("860x540")
        self.root.resizable(False, False)
        self.root.configure(bg='#0d1117')
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x  = (sw - 860) // 2
        y  = (sh - 540) // 2
        self.root.geometry(f"860x540+{x}+{y}")

        BG     = '#0d1117'
        PANEL  = '#161b22'
        BLUE   = '#2d7dd2'
        BLUE2  = '#1a5fa8'
        CYAN   = '#58c4dd'
        FG     = '#e6edf3'
        FG2    = '#8b949e'
        ACCENT = '#21262d'

        self.canvas = tk.Canvas(self.root, width=860, height=540,
                                bg=BG, highlightthickness=0)
        self.canvas.pack(fill='both', expand=True)

        # Background grid
        for gx in range(0, 860, 40):
            self.canvas.create_line(gx, 0, gx, 540, fill='#161b22', width=1)
        for gy in range(0, 540, 40):
            self.canvas.create_line(0, gy, 860, gy, fill='#161b22', width=1)

        # Left accent bar
        self.canvas.create_rectangle(0, 0, 5, 540, fill=BLUE, outline='')

        # Org badge top-left
        self.canvas.create_text(28, 22, text='-iv \u00b7  IV',
                                fill=FG2, font=('Segoe UI', 9, 'bold'), anchor='w')
        self.canvas.create_line(28, 34, 220, 34, fill=ACCENT, width=1)

        # Bottom status bar
        self.canvas.create_rectangle(0, 510, 860, 540, fill=PANEL, outline='')
        self.canvas.create_line(0, 510, 860, 510, fill=BLUE2, width=1)
        self._status_id = self.canvas.create_text(
            28, 525, text='Initialising system\u2026',
            fill=FG2, font=('Consolas', 9), anchor='w')
        self.canvas.create_text(
            832, 525, text='v3.0  |  BUILD 2025',
            fill=FG2, font=('Consolas', 9), anchor='e')

        # Hexagon logo
        cx, cy, r = 170, 265, 68
        import math as _math
        pts_outer = []
        for i in range(6):
            a = _math.radians(60 * i - 30)
            pts_outer += [cx + r * _math.cos(a), cy + r * _math.sin(a)]
        self.canvas.create_polygon(pts_outer, outline=BLUE, fill='#0d1a2a', width=2)
        pts_inner = []
        ri = 48
        for i in range(6):
            a = _math.radians(60 * i - 30)
            pts_inner += [cx + ri * _math.cos(a), cy + ri * _math.sin(a)]
        self.canvas.create_polygon(pts_inner, outline=CYAN, fill='#071020', width=1)
        self.canvas.create_text(cx, cy, text='\u2b21',
                                fill=BLUE, font=('Segoe UI', 42))

        # Corner brackets around hex
        br_off, br_len, br_w = 84, 18, 2
        for bx, by, xs, ys in [(cx - br_off, cy - br_off, 1, 1),
                                (cx + br_off, cy - br_off, -1, 1),
                                (cx + br_off, cy + br_off, -1, -1),
                                (cx - br_off, cy + br_off, 1, -1)]:
            self.canvas.create_line(bx, by, bx + xs * br_len, by,
                                    fill=CYAN, width=br_w)
            self.canvas.create_line(bx, by, bx, by + ys * br_len,
                                    fill=CYAN, width=br_w)

        # Title block
        self.canvas.create_text(278, 222,
                                text='INDICATOR DETECTION SYSTEM',
                                fill=FG, font=('Segoe UI', 22, 'bold'), anchor='w')
        self.canvas.create_text(278, 254,
                                text='Advanced Image Analysis Pipeline',
                                fill=CYAN, font=('Segoe UI', 11), anchor='w')
        self.canvas.create_line(278, 272, 835, 272, fill=BLUE2, width=1)

        # Spec rows
        specs = [
            ('MODULE',   'Watershed Segmentation  \u00b7  Scikit-Image'),
            ('PIPELINE', '11-Step Analysis  \u00b7  ROI + Batch Processing'),
            ('OUTPUT',   'Indicator Classification  \u00b7  Perspective Warp'),
            ('PLATFORM', 'Python  \u00b7  OpenCV  \u00b7  Tkinter  \u00b7  NumPy'),
        ]
        sy = 292
        for lbl, val in specs:
            self.canvas.create_text(278, sy, text=lbl,
                                    fill=BLUE, font=('Consolas', 8, 'bold'), anchor='w')
            self.canvas.create_text(368, sy, text=val,
                                    fill=FG2, font=('Consolas', 8), anchor='w')
            sy += 18
       

        # Progress bar
        bx1, by1, bx2, by2 = 278, 410, 835, 424
        self.canvas.create_rectangle(bx1, by1, bx2, by2,
                                     fill=ACCENT, outline=BLUE2, width=1)
        self._bar = self.canvas.create_rectangle(
            bx1 + 1, by1 + 1, bx1 + 1, by2 - 1, fill=BLUE, outline='')
        self._bar_x1    = bx1 + 1
        self._bar_width = (bx2 - 1) - (bx1 + 1)
        self._pct_id = self.canvas.create_text(
            bx1, by1 - 10, text='0%',
            fill=CYAN, font=('Consolas', 8, 'bold'), anchor='w')

        # Animation state
        self._progress = 0
        self._msg_idx  = 0
        self._step_msgs = [
            (0,   'Initialising system\u2026'),
            (12,  'Loading image processing modules\u2026'),
            (28,  'Configuring watershed pipeline\u2026'),
            (45,  'Preparing ROI segmentation engine\u2026'),
            (62,  'Loading batch processing routines\u2026'),
            (78,  'Calibrating indicator classifier\u2026'),
            (90,  'Finalising perspective warp module\u2026'),
            (100, 'System ready.'),
        ]
        self._animate()

    def _animate(self):
        if self._progress >= 100:
            self._finish()
            return
        self._progress = min(self._progress + 1, 100)
        pct = self._progress
        new_right = self._bar_x1 + int(self._bar_width * pct / 100)
        coords = self.canvas.coords(self._bar)
        self.canvas.coords(self._bar, coords[0], coords[1], new_right, coords[3])
        self.canvas.itemconfig(self._pct_id, text=f'{pct}%')
        while (self._msg_idx < len(self._step_msgs) - 1 and
               pct >= self._step_msgs[self._msg_idx + 1][0]):
            self._msg_idx += 1
        self.canvas.itemconfig(self._status_id,
                               text=self._step_msgs[self._msg_idx][1])
        delay = 18 if pct < 95 else 40
        self.root.after(delay, self._animate)

    def _finish(self):
        BLUE  = '#2d7dd2'
        BLUE2 = '#1a5fa8'
        FG    = '#e6edf3'
        btn_frame = tk.Frame(self.root, bg='#161b22', bd=0)
        self.canvas.create_window(557, 460, window=btn_frame)
        tk.Button(btn_frame,
                  text='  ENTER SYSTEM  \u203a',
                  command=self._open_login,
                  bg=BLUE, fg=FG, relief='flat',
                  font=('Segoe UI', 11, 'bold'),
                  padx=28, pady=8,
                  activebackground=BLUE2,
                  activeforeground='#ffffff',
                  cursor='hand2', bd=0).pack()
        self.root.bind('<Return>', lambda e: self._open_login())

    def _open_login(self):
        self.root.destroy()
        _launch_login()


def main():
    _launch_main()


if __name__ == "__main__":
    main()

#################################   The terminal cleanly shuts down the Tkinter application by closing its main window.      ################################################


import signal
import sys
import tkinter as tk

# ------------------------------------------------------------------
# 2.  Signal handler – called when the process receives SIGINT
# ------------------------------------------------------------------
'''def _sigint_handler(signum, frame):
    # Queue the destroy operation on the Tk event loop
    root.after(0, root.destroy)

# Install the handler **before** we start the event loop
signal.signal(signal.SIGINT, _sigint_handler)

# ------------------------------------------------------------------
# 3.  Run the Tkinter event loop
# ------------------------------------------------------------------
try:
    root.mainloop()
except KeyboardInterrupt:

    root.destroy()

# ------------------------------------------------------------------
# 4.  Exit with a clean status code
# ------------------------------------------------------------------
sys.exit(0)'''

'''The terminal sends a SIGINT (interrupt) signal to the running program.
The program’s signal‑handler receives that signal.
Inside the handler, the code tells Tkinter to close its main window (root.destroy()).
Tkinter’s event loop sees that request, shuts the window down, and then stops running.'''


# ------------------------------------------------------------------
# 2.  Signal handler – called when the process receives SIGINT
# ------------------------------------------------------------------
def _sigint_handler(signum, frame):
    """Called when the user presses Ctrl‑C in the terminal."""
    # Schedule the quit in the GUI thread
    root.after(0, root.quit)  # stop the event loop ASAP


# Install the handler before the event loop starts
signal.signal(signal.SIGINT, _sigint_handler)

# ------------------------------------------------------------------
# 3.  Schedule automatic shutdown after 10 seconds
# ------------------------------------------------------------------
root.after(10_000, root.quit)  # 10 000 ms = 10 s

# ------------------------------------------------------------------
# 4.  Run the Tkinter event loop
# ------------------------------------------------------------------
try:
    root.mainloop()  # will return immediately after root.quit()
except KeyboardInterrupt:
    # Fallback for environments that don’t deliver SIGINT
    root.quit()

# ------------------------------------------------------------------
# 5.  Close the window after the loop has exited
# ------------------------------------------------------------------
root.destroy()  # safe to call now – the loop is finished

# ------------------------------------------------------------------
# 6.  Exit with a clean status code
# ------------------------------------------------------------------
sys.exit(0)

    
    

